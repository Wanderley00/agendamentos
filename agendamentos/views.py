# agendamentos/views.py

from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden
import json
import re
from django.db import transaction
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from datetime import datetime, time, timedelta
import calendar
from django.core.exceptions import ValidationError
import colorsys
from django.utils.text import slugify
from django.db.models import Sum, Count
from django.db.models import Sum, Count, OuterRef, Subquery, Q  # << ADICIONE Q
from datetime import datetime, time, timedelta  # << ADICIONE timedelta
import calendar  # << ADICIONE calendar
from django.conf import settings
from django.utils import timezone
from datetime import timedelta
from .mercadopago_service import MercadoPagoService
import logging
import re  # Para limpar o telefone
from django.db.models import Q
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string

from .webhooks import disparar_notificacao

from django.http import JsonResponse
from django.utils import timezone
from datetime import timedelta
from django.views.decorators.csrf import csrf_exempt
from .models import Agendamento
import re
import os

# --- IMPORTAÇÃO CORRIGIDA ---
# Adicionamos EmpreendedorProfile e removemos importações duplicadas
from .models import (
    Servico, Agendamento, Cliente, Negocio, EmpreendedorProfile, Despesa, HorarioTrabalho, Aviso, DiaBloqueado,
    Categoria, PrecoManutencao, DespesaRecorrente
)

from django.db import transaction  # Adicione esta importação
from django.utils.crypto import get_random_string  # Adicione esta importação

logger = logging.getLogger(__name__)

# ---
# Views do Cliente (A maior parte já estava correta!)
# ---


def index(request, empreendedor_slug):
    try:
        # 1. Tenta encontrar o negócio pelo slug
        negocio = Negocio.objects.get(slug=empreendedor_slug)

        # 2. Se encontrar, renderiza o SPA (index.html)
        return render(request, 'agendamentos/index.html')

    except Negocio.DoesNotExist:
        # 3. Se NÃO encontrar, renderiza uma página de erro personalizada
        context = {'slug_invalido': empreendedor_slug}
        return render(request, 'agendamentos/negocio_nao_encontrado.html', context, status=404)


def lista_servicos(request, empreendedor_slug):
    # --- LÓGICA INTEIRA MODIFICADA ---
    negocio = get_object_or_404(Negocio, slug=empreendedor_slug)

    # Busca todos os serviços principais (exclui os que são SÓ manutenção, se houver)
    # Por enquanto, nosso modelo assume que todo serviço "pai" é agendável
    servicos_principais = Servico.objects.filter(
        negocio=negocio
    ).prefetch_related(
        'precos_manutencao'  # Puxa os tiers de manutenção
    )

    cliente = None
    if request.user.is_authenticated:
        try:
            cliente = Cliente.objects.get(user=request.user, negocio=negocio)
        except Cliente.DoesNotExist:
            # Usuário logado mas não é cliente (ex: admin de outro negócio)
            cliente = None

    # Encontra o último agendamento CONCLUÍDO (ou Confirmado) do cliente
    ultimo_agendamento = None
    if cliente:
        # Subquery para encontrar a data mais recente
        # (Isso garante que estamos pegando o mais recente MESMO)
        data_recente = Agendamento.objects.filter(
            cliente=cliente,
            servico__categoria=OuterRef('servico__categoria'),
            status__in=['Concluído', 'Confirmado']
        ).order_by('-data', '-horario').values('data')[:1]

        # Agora busca o agendamento naquela data
        # (Necessário porque pode haver múltiplos no mesmo dia)
        ultimo_agendamento_por_categoria = Agendamento.objects.filter(
            cliente=cliente,
            servico__categoria=OuterRef('servico__categoria'),
            status__in=['Concluído', 'Confirmado'],
            data=Subquery(data_recente)
        ).order_by('-horario').values('id')[:1]

        # Pega o último agendamento geral do cliente
        ultimo_agendamento = Agendamento.objects.filter(
            cliente=cliente,
            status__in=['Concluído', 'Confirmado']
        ).order_by('-data', '-horario').first()

    data_final_servicos = []

    for servico in servicos_principais:
        # Dados base do serviço (preço cheio)
        servico_data = {
            'id': servico.id,
            'name': servico.nome,
            'description': servico.descricao,
            'duracao_minutos': servico.duracao_minutos,
            'duracao_formatada': servico.duracao_formatada,
            'price': float(servico.preco),
            'icon': '✨',
            'image_url': servico.imagem.url if servico.imagem else None,
            'categoria_id': servico.categoria_id,
            'tiers_manutencao': []  # Lista de manutenções disponíveis
        }

        # Se o cliente não está logado, ou o serviço não tem categoria,
        # ou não há um último agendamento, mostramos apenas o preço cheio.
        if not cliente or not servico.categoria or not ultimo_agendamento:
            data_final_servicos.append(servico_data)
            continue  # Próximo serviço

        # O cliente está logado e o serviço tem categoria.
        # Verificamos se o ÚLTIMO agendamento dele foi dessa categoria.
        if (ultimo_agendamento.servico.categoria_id != servico.categoria_id):
            # Cenário A: Categoria diferente (Ex: Unha -> Cílios).
            # Isso é uma "troca de serviço", mostra o preço cheio.
            data_final_servicos.append(servico_data)
            continue  # Próximo serviço

        # SE CHEGOU AQUI:
        # A CATEGORIA É A MESMA.
        # Agora, verificamos se é o *mesmo serviço* ou uma *troca* dentro da categoria.

        # Esta é a nova variável crucial:
        is_troca_servico = (ultimo_agendamento.servico.id != servico.id)

        # Calcula há quantos dias foi o último serviço
        dias_desde_ultimo_servico = (
            datetime.now().date() - ultimo_agendamento.data).days

        # --- INÍCIO DA NOVA LÓGICA ---
        # Decide qual texto será usado no pop-up do frontend
        motivo_preco_cheio = ""
        if is_troca_servico:
            motivo_preco_cheio = "Troca de serviço"
        else:
            # Se não é troca, o único outro motivo para o preço cheio é expiração
            motivo_preco_cheio = "Expiração das manutenções"

        # Adiciona o motivo ao dicionário que vai para o frontend
        servico_data['motivo_preco_cheio'] = motivo_preco_cheio
        # --- FIM DA NOVA LÓGICA ---

        tiers_disponiveis = []

        for tier in servico.precos_manutencao.all():

            # --- ESTA É A MUDANÇA PRINCIPAL ---
            # O tier SÓ pode estar ativo se:
            # 1. NÃO for uma troca de serviço E
            # 2. Estiver dentro do range de dias.
            is_active = (
                not is_troca_servico and  # <-- ADICIONADO
                (tier.dias_min <= dias_desde_ultimo_servico <= tier.dias_max)
            )
            # --- FIM DA MUDANÇA ---

            # --- INÍCIO DA NOVA LÓGICA ---
            inactivity_message = ""
            if not is_active:
                if is_troca_servico:
                    # Este é o motivo se for uma troca de serviço
                    inactivity_message = "Indisponível para troca de serviço."
                elif dias_desde_ultimo_servico < tier.dias_min:
                    # Este é o motivo se AINDA NÃO CHEGOU O TEMPO
                    inactivity_message = f"Este período de manutenção estará disponível em {tier.dias_min} dias."
                elif dias_desde_ultimo_servico > tier.dias_max:
                    # Este é o motivo se JÁ PASSOU O TEMPO
                    inactivity_message = "Este período de manutenção já passou."
                else:
                    # Fallback
                    inactivity_message = "Este período não está disponível."
            # --- FIM DA NOVA LÓGICA ---

            tiers_disponiveis.append({
                'id': tier.id,
                'nome_tier': tier.nome_tier,
                'preco': float(tier.preco),
                'duracao_minutos': tier.duracao_minutos,
                'duracao_formatada': tier.duracao_formatada,
                'dias_min': tier.dias_min,
                'dias_max': tier.dias_max,
                'is_active': is_active,  # Agora 'is_active' será false se for troca
                'inactivity_message': inactivity_message
            })

        servico_data['tiers_manutencao'] = tiers_disponiveis
        data_final_servicos.append(servico_data)

    # A resposta da API agora é padronizada
    data = {
        'cor_primaria': negocio.cor_primaria,
        'servicos': data_final_servicos
    }
    return JsonResponse(data)
    # --- FIM DA LÓGICA MODIFICADA ---


def get_profissionais_por_servico(request, empreendedor_slug, servico_id):
    try:
        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)
        servico = get_object_or_404(Servico, id=servico_id, negocio=negocio)

        profissionais = servico.profissionais_que_executam.all()

        data = [
            {
                'id': prof.id,
                'nome': prof.user.get_full_name() or prof.user.username,
                'foto_url': prof.foto.url if prof.foto else None,
                # Futuramente você pode adicionar foto_url, etc.
            } for prof in profissionais
        ]

        # Se não houver profissionais, mas o serviço existir, retorna lista vazia
        return JsonResponse(data, safe=False)

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def register_user(request, empreendedor_slug):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        phone = data.get('phone')  # <-- Pega o telefone
        password = get_random_string(length=14)

        try:
            negocio = Negocio.objects.get(slug=empreendedor_slug)
        except Negocio.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Negócio não encontrado.'}, status=404)

        # 1. Verificação de Email (já existente)
        if User.objects.filter(username=email).exists():
            return JsonResponse({'status': 'error', 'message': 'Este e-mail já está em uso.'}, status=400)

        # --- ADICIONE ESTE BLOCO ---
        # 2. Verificação de Telefone
        if not phone:
            return JsonResponse({'status': 'error', 'message': 'O telefone é obrigatório.'}, status=400)

        telefone_limpo = re.sub(r'\D', '', phone)  # Este será o novo username

        # 1. Verificação de Telefone (que agora é o USERNAME)
        # Verifica se o telefone já existe GLOBALMENTE como um username
        if User.objects.filter(username=telefone_limpo).exists():
            return JsonResponse({'status': 'error', 'message': 'Este telefone já está em uso.'}, status=400)

        # 2. Verificação de Email (SÓ SE ele foi fornecido)
        if email:
            if User.objects.filter(email=email).exists():
                return JsonResponse({'status': 'error', 'message': 'Este e-mail já está em uso.'}, status=400)
        else:
            email = None  # Garante que seja None se for "" (vazio)

        # 3. Verificação de Telefone no Negócio (validação do Cliente)
        if Cliente.objects.filter(negocio=negocio, telefone=telefone_limpo).exists():
            return JsonResponse({'status': 'error', 'message': 'Este telefone já está cadastrado neste negócio.'}, status=400)
        # --- FIM DA ADIÇÃO ---

        user = User.objects.create_user(
            username=telefone_limpo,  # <-- CORREÇÃO (usando telefone)
            email=email,
            password=password,
            first_name=data.get('name'),
            last_name=data.get('lastname')
        )

        data_nascimento = data.get('nascimento')
        # --- MUDANÇA AQUI ---
        if not data_nascimento:
            # Se não houver data de nascimento, retorna um erro
            user.delete()  # Exclui o usuário que foi criado
            return JsonResponse({'status': 'error', 'message': 'A data de nascimento é obrigatória para o cadastro.'}, status=400)

        Cliente.objects.create(
            user=user,
            telefone=telefone_limpo,
            negocio=negocio,
            data_nascimento=data_nascimento
        )

        login(request, user)
        return JsonResponse({'status': 'success', 'message': 'Registro e login bem-sucedidos!'}, status=201)
    return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)


def login_user(request, empreendedor_slug):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')

        try:
            negocio = Negocio.objects.get(slug=empreendedor_slug)
        except Negocio.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Negócio não encontrado.'}, status=404)

        user = authenticate(request, username=email, password=password)

        if user is not None:
            try:
                cliente = Cliente.objects.get(user=user, negocio=negocio)
                login(request, user)
                return JsonResponse({'status': 'success', 'message': 'Login bem-sucedido!'})
            except Cliente.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'E-mail ou senha inválidos.'}, status=401)
        else:
            return JsonResponse({'status': 'error', 'message': 'E-mail ou senha inválidos.'}, status=401)
    return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)


def login_user_with_phone(request, empreendedor_slug):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)

    try:
        data = json.loads(request.body)
        phone = data.get('phone')
        nascimento = data.get('nascimento')  # Espera YYYY-MM-DD

        if not phone or not nascimento:
            return JsonResponse({'status': 'error', 'message': 'Telefone e data de nascimento são obrigatórios.'}, status=400)

        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)

        # --- Lógica de Autenticação ---
        # 1. Limpa o telefone (caso o JS falhe)
        telefone_limpo = re.sub(r'\D', '', phone)

        # 2. Busca o cliente
        cliente = Cliente.objects.filter(
            negocio=negocio,
            telefone=telefone_limpo,
            data_nascimento=nascimento
        ).select_related('user').first()  # Traz o 'user' junto na query

        if cliente and cliente.user:
            # 3. Se encontrou, loga o usuário associado
            login(request, cliente.user)
            return JsonResponse({'status': 'success', 'message': 'Login bem-sucedido!'})
        else:
            # 4. Se não encontrou, retorna erro
            return JsonResponse({'status': 'error', 'message': 'Telefone ou data de nascimento inválidos.'}, status=401)

    except (Negocio.DoesNotExist):
        return JsonResponse({'status': 'error', 'message': 'Negócio não encontrado.'}, status=404)
    except Exception as e:
        # Pega outros erros (ex: data de nascimento em formato inválido)
        return JsonResponse({'status': 'error', 'message': f'Erro no servidor: {str(e)}'}, status=500)


def logout_user(request, empreendedor_slug):
    logout(request)
    return JsonResponse({'status': 'success', 'message': 'Logout bem-sucedido!'})


def check_auth_status(request, empreendedor_slug):
    if request.user.is_authenticated:
        return JsonResponse({
            'isAuthenticated': True,
            'user': {
                'name': request.user.first_name,
                'lastname': request.user.last_name,
                'email': request.user.email
            }
        })
    else:
        return JsonResponse({'isAuthenticated': False})


# @login_required(login_url=None)
def lista_meus_agendamentos(request, empreendedor_slug):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'Autenticação necessária.'}, status=403)
    try:
        cliente = request.user.cliente
        agendamentos = Agendamento.objects.filter(cliente=cliente).select_related(
            'servico').order_by('-data', '-horario')

        data = []
        for agendamento in agendamentos:
            appointment_datetime = timezone.make_aware(
                datetime.combine(agendamento.data, agendamento.horario))
            now = timezone.now()
            time_difference = appointment_datetime - now

            # --- MUDANÇA AQUI ---
            # Verifica o tempo E o status
            can_reschedule = (time_difference > timedelta(hours=24)) and (
                agendamento.status in ['Confirmado', 'Pendente'])

            # --- MUDANÇA 1: Buscar a Imagem ---
            service_image_url = None
            if agendamento.servico.imagem:
                service_image_url = agendamento.servico.imagem.url

            # --- MUDANÇA 2: Buscar o Profissional ---
            profissional_nome = "Profissional não definido"  # Fallback
            if agendamento.empreendedor_executor:
                profissional_nome = agendamento.empreendedor_executor.user.get_full_name(
                ) or agendamento.empreendedor_executor.user.username

            data.append({
                'id': agendamento.id,
                'service': agendamento.servico.nome,
                'serviceId': agendamento.servico.id,
                'date': agendamento.data.strftime('%Y-%m-%d'),
                'time': agendamento.horario.strftime('%H:%M'),
                'status': agendamento.status,
                'can_reschedule': can_reschedule,

                # --- CAMPOS ADICIONADOS PARA O NOVO CARD ---
                'serviceImageUrl': service_image_url,
                'profissional': profissional_nome,
                # Usa o preço final salvo
                'preco': str(agendamento.preco_final)
            })
        return JsonResponse(data, safe=False)
    except Cliente.DoesNotExist:
        return JsonResponse([], safe=False)


@login_required(login_url=None)
def cancelar_agendamento(request, agendamento_id, empreendedor_slug):
    if request.method == 'POST':
        try:
            agendamento = Agendamento.objects.get(
                id=agendamento_id, cliente=request.user.cliente)
            agendamento.delete()
            return JsonResponse({'status': 'success', 'message': 'Agendamento cancelado com sucesso.'})
        except Agendamento.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Agendamento não encontrado ou não pertence a você.'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)


@csrf_exempt
@login_required(login_url=None)
def criar_agendamento(request, empreendedor_slug):
    """
    Cria um agendamento.
    Se o adiantamento > 0, status = 'Aguardando Pagamento' e gera PIX.
    Se o adiantamento == 0, status = 'Confirmado'.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)

    try:
        data = json.loads(request.body)
        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)
        cliente = request.user.cliente
        servico = Servico.objects.get(id=data['serviceId'], negocio=negocio)
        profissional = EmpreendedorProfile.objects.get(
            id=data['empreendedorId'], negocio=negocio)

        if cliente.negocio != negocio:
            return JsonResponse({'status': 'error', 'message': 'Erro de permissão.'}, status=403)

        # 1. Verifica se o horário ainda está disponível (SEGURANÇA)
        # ... (Omissão da lógica de verificação de horário para brevidade, mantendo a existente) ...

        # 2. Lógica de Manutenção
        tier_manutencao = None
        tier_id = data.get('tierManutencaoId')
        if tier_id:
            tier_manutencao = PrecoManutencao.objects.get(
                id=tier_id, servico_pai=servico)

        # 3. Cria o Agendamento (instancia o objeto)
        ag = Agendamento(
            cliente=cliente,
            servico=servico,
            data=data['date'],
            horario=data['time'],
            empreendedor_executor=profissional,
            tier_manutencao=tier_manutencao,
            status='Pendente',
            # Pega observações do usuário se houver
            observacoes=data.get('observacoes', '')
        )

        # 4. Roda o .save() INICIAL para calcular preco_final e valor_adiantamento
        ag.save()

        # --- INÍCIO DA NOVA ADIÇÃO: Inserir detalhes financeiros nas observações ---
        # Agora que o .save() rodou, temos os valores calculados no objeto 'ag'
        if ag.valor_adiantamento > 0:
            restante = ag.preco_final - ag.valor_adiantamento
            # Calcula porcentagem real baseada nos valores finais
            porcentagem = int((ag.valor_adiantamento / ag.preco_final) * 100)

            detalhe_financeiro = (
                f"\n[SISTEMA] Resumo Financeiro:\n"
                f"- Total: R$ {ag.preco_final}\n"
                f"- Adiantamento ({porcentagem}%): R$ {ag.valor_adiantamento} (Pago Online)\n"
                f"- Restante a pagar no dia: R$ {restante}"
            )

            # Concatena com o que já existia
            if ag.observacoes:
                ag.observacoes += "\n" + detalhe_financeiro
            else:
                ag.observacoes = detalhe_financeiro
        # --- FIM DA NOVA ADIÇÃO ---

        # 5. Define o status de pagamento
        if negocio.pagamento_online_habilitado and ag.valor_adiantamento > 0:
            ag.status_pagamento = 'Aguardando Pagamento'
        else:
            ag.status_pagamento = 'Pendente'

        # Salva novamente com as observações atualizadas e status
        ag.save()

        # 6. Verifica se o pagamento (PIX) é necessário
        if ag.status_pagamento == 'Pendente':
            logger.info(
                f"Agendamento {ag.id} criado como Pendente (sem adiantamento).")
            return JsonResponse({
                'status': 'success',
                'payment_required': False,
                'agendamento_id': ag.id
            }, status=201)

        # 7. Pagamento PIX é necessário
        try:
            # Obter token do profissional
            access_token = profissional.get_access_token()
            if not access_token:
                raise Exception(
                    "Profissional não possui conta Mercado Pago vinculada.")

            # Inicializa o serviço com o token do profissional
            mp = MercadoPagoService(access_token=access_token)
            payment_data = mp.criar_pagamento_pix(ag)

            if not payment_data:
                raise Exception("Falha ao gerar PIX no Mercado Pago.")

            # Atualiza o agendamento com os dados do PIX
            ag.payment_id_mp = payment_data["payment_id"]
            ag.payment_qrcode = payment_data["qr_code"]
            ag.payment_qrcode_image = payment_data["qr_code_base64"]
            ag.payment_expires = payment_data["expires_at"]
            ag.save()

            logger.info(
                f"Agendamento {ag.id} aguardando pagamento (PIX gerado).")

            return JsonResponse({
                'status': 'pending_payment',
                'payment_required': True,
                'agendamento_id': ag.id,
                'payment_id_mp': ag.payment_id_mp,
                'qr_code': ag.payment_qrcode,
                'qr_code_base64': ag.payment_qrcode_image,
                'expires_at': ag.payment_expires.isoformat()
            }, status=201)

        except Exception as e:
            logger.error(
                f"Falha na API do MP para Agendamento {ag.id}. Cancelando. Erro: {e}")
            ag.status = 'Cancelado'
            # Adiciona erro nas observações
            error_msg = f'\n[ERRO SISTEMA] Falha ao gerar PIX: {e}'
            if ag.observacoes:
                ag.observacoes += error_msg
            else:
                ag.observacoes = error_msg
            ag.save()
            return JsonResponse({'status': 'error', 'message': f'Erro ao processar pagamento: {e}'}, status=500)

    except (Servico.DoesNotExist, EmpreendedorProfile.DoesNotExist, Cliente.DoesNotExist, PrecoManutencao.DoesNotExist):
        return JsonResponse({'status': 'error', 'message': 'Dados inválidos.'}, status=404)
    except Exception as e:
        logger.error(
            f"Erro inesperado em criar_agendamento: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@csrf_exempt
def mercadopago_webhook(request):
    """
    Recebe notificações de pagamento do Mercado Pago.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)

    try:
        data = json.loads(request.body)
        logger.info(f"Webhook Mercado Pago recebido: {data}")

        if data.get("type") == "payment":
            payment_id_mp = str(data.get("data", {}).get("id"))
            if not payment_id_mp:
                return JsonResponse({'status': 'ignorado', 'message': 'Sem ID de pagamento.'}, status=200)

            logger.info(
                f"Processando notificação para Payment ID: {payment_id_mp}")

            # 1. Tenta encontrar o agendamento pelo ID do Pagamento do MP
            try:
                agendamento = Agendamento.objects.get(
                    payment_id_mp=payment_id_mp)
            except Agendamento.DoesNotExist:
                logger.warning(
                    f"Webhook para Payment ID {payment_id_mp} não encontrado no banco de dados.")
                return JsonResponse({'status': 'nao_encontrado'}, status=200)

            # 2. Se o agendamento já foi processado, ignora
            if agendamento.status_pagamento != 'Aguardando Pagamento':
                logger.info(
                    f"Agendamento {agendamento.id} já processado (Status Pag: {agendamento.status_pagamento}). Ignorando webhook.")
                return JsonResponse({'status': 'ja_processado'}, status=200)

            # 3. Verifica o status real do pagamento na API do MP
            # --- MODIFICAÇÃO: Obter token do profissional executor deste agendamento ---
            executor = agendamento.empreendedor_executor
            access_token = executor.get_access_token() if executor else None

            if not access_token:
                logger.critical(
                    f"Executor do agendamento {agendamento.id} não possui token MP. Falha na verificação.")
                return JsonResponse({'status': 'error', 'message': 'Token não encontrado'}, status=500)

            # Inicializa o serviço com o token específico deste profissional
            mp = MercadoPagoService(access_token=access_token)
            status_real = mp.verificar_status_pagamento(payment_id_mp)

            if status_real == "approved":
                agendamento.status = 'Pendente'

                if agendamento.valor_adiantamento < agendamento.preco_final:
                    agendamento.status_pagamento = 'Adiantamento Realizado'
                    agendamento.observacoes = f"Adiantamento {payment_id_mp} aprovado via webhook. Aguardando confirmação manual."
                else:
                    agendamento.status_pagamento = 'Pago'
                    agendamento.observacoes = f"Pagamento integral {payment_id_mp} aprovado via webhook. Aguardando confirmação manual."

                agendamento.save()

                # 1. Limpa tudo que não é número
                telefone_limpo = re.sub(
                    r'\D', '', agendamento.cliente.telefone)

                # 2. Se for um número brasileiro sem DDI (11 dígitos), adiciona o 55
                if len(telefone_limpo) == 11:
                    telefone_limpo = f"55{telefone_limpo}"

                # --- AQUI ENTRA O N8N ---
                try:
                    dados_notificacao = {
                        "cliente_nome": agendamento.cliente.user.get_full_name() or agendamento.cliente.user.username,
                        # Formato: 5511999999999
                        "cliente_telefone": telefone_limpo,
                        "servico_nome": agendamento.servico.nome,
                        "profissional": agendamento.empreendedor_executor.user.get_full_name(),
                        "data": agendamento.data.strftime('%d/%m/%Y'),
                        "horario": agendamento.horario.strftime('%H:%M'),
                        "local_nome": agendamento.servico.negocio.nome_negocio,
                        "valor": float(agendamento.servico.preco),
                        "link_google_calendar": "https://calendar.google.com/..."  # Podemos gerar isso depois
                    }

                    disparar_notificacao(
                        'pagamento_confirmado', dados_notificacao)
                    print("🚀 Notificação enviada para o n8n!")

                except Exception as e:
                    print(f"Erro ao preparar dados n8n: {e}")

                logger.info(
                    f"PAGAMENTO APROVADO: Agendamento {agendamento.id} PAGO. Aguardando confirmação manual.")

            elif status_real in ["rejected", "cancelled", "expired"]:
                agendamento.status = 'Cancelado'
                agendamento.status_pagamento = 'Cancelado'
                agendamento.observacoes = f"Pagamento {payment_id_mp} falhou ou expirou (Status: {status_real})."
                agendamento.save()
                logger.warning(
                    f"PAGAMENTO FALHOU: Agendamento {agendamento.id} cancelado.")

            else:
                logger.info(
                    f"Status '{status_real}' recebido para Agendamento {agendamento.id}. Nenhuma ação tomada.")

    except json.JSONDecodeError:
        logger.error("Erro ao decodificar JSON do webhook.")
        return JsonResponse({'status': 'error', 'message': 'JSON inválido.'}, status=400)
    except Exception as e:
        logger.error(f"Erro inesperado no webhook: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': 'Erro interno.'}, status=500)

    return JsonResponse({"status": "recebido"}, status=200)


@login_required(login_url=None)
def check_booking_status(request, agendamento_id, empreendedor_slug):
    """
    Verifica o status de um agendamento no banco de dados.
    Usado pelo frontend (polling) para atualizar a UI após o pagamento.
    """
    try:
        agendamento = get_object_or_404(
            Agendamento,
            id=agendamento_id,
            cliente=request.user.cliente
        )

        # Retorna o status atual do agendamento
        return JsonResponse({
            # Ex: "Pendente", "Confirmado", "Cancelado"
            'status': agendamento.status,
            # --- NOVA ADIÇÃO ---
            # Ex: "Aguardando Pagamento", "Adiantamento Realizado", "Pago"
            'status_pagamento': agendamento.status_pagamento,
            # --- FIM DA ADIÇÃO ---
            'agendamento_id': agendamento.id
        })

    except Agendamento.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Agendamento não encontrado.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def get_horarios_disponiveis(request, empreendedor_slug):
    # --- LÓGICA MODIFICADA ---
    data_str = request.GET.get('data')
    servico_id = request.GET.get('servico_id')
    empreendedor_id = request.GET.get('empreendedor_id')
    # --- NOVA ADIÇÃO ---
    # O JS deve enviar a duração EXATA do que foi selecionado
    # (seja o serviço principal ou o tier de manutenção)
    duracao_selecionada = request.GET.get('duracao')
    # --- FIM DA ADIÇÃO ---

    if not data_str or not servico_id or not empreendedor_id or not duracao_selecionada:
        return JsonResponse({'status': 'error', 'message': 'Data, serviço, profissional e duração são obrigatórios.'}, status=400)

    try:
        data = datetime.strptime(data_str, '%Y-%m-%d').date()
        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)
        # Valida o serviço e o profissional
        servico = Servico.objects.get(id=servico_id, negocio=negocio)
        profissional = EmpreendedorProfile.objects.get(
            id=empreendedor_id, negocio=negocio)
        # --- NOVA ADIÇÃO ---
        duracao_novo_servico = timedelta(minutes=int(duracao_selecionada))
        # --- FIM DA ADIÇÃO ---

    except (ValueError, Servico.DoesNotExist, EmpreendedorProfile.DoesNotExist):
        return JsonResponse({'status': 'error', 'message': 'Data, serviço ou profissional inválido.'}, status=400)

    if DiaBloqueado.objects.filter(empreendedor=profissional, data=data).exists():
        return JsonResponse([], safe=False)

    # --- INÍCIO DA NOVA LÓGICA DE HORÁRIOS ---

    # 1. Obter o dia da semana (0=Segunda, 1=Terça, ..., 6=Domingo)
    dia_da_semana = data.weekday()

    # 2. Buscar os blocos de trabalho personalizados deste profissional para este dia
    blocos_de_trabalho = HorarioTrabalho.objects.filter(
        empreendedor=profissional,
        dia_da_semana=dia_da_semana
    ).order_by('hora_inicio')

    # 3. Definir o intervalo (ex: de 30 em 30 min)
    intervalo_minutos = 30  # (Isso deve ser configurável no futuro)

    # --- MODIFICADO ---
    # Não usamos mais servico.duracao_minutos
    # duracao_novo_servico = timedelta(minutes=servico.duracao_minutos)
    # --- FIM DA MODIFICAÇÃO ---

    # --- MODIFICADO ---
    # Busca agendamentos do dia E USA A DURAÇÃO FINAL
    agendamentos_do_dia = Agendamento.objects.filter(
        data=data,
        empreendedor_executor=profissional,
        status__in=['Confirmado', 'Pendente']  # Não conta cancelados
    )

    blocos_ocupados = []
    for agendamento in agendamentos_do_dia:
        inicio_naive = datetime.combine(data, agendamento.horario)
        inicio = timezone.make_aware(inicio_naive)
        # USA A DURAÇÃO REAL DO AGENDAMENTO
        duracao_agendamento = agendamento.duracao_final
        fim = inicio + timedelta(minutes=duracao_agendamento)
        blocos_ocupados.append((inicio, fim))
    # --- FIM DA MODIFICAÇÃO ---

    # --- LÓGICA DE VERIFICAÇÃO DE CONFLITO (existente) ---
    def verificar_conflito(inicio_potencial, fim_potencial):
        for inicio_ocupado, fim_ocupado in blocos_ocupados:
            if max(inicio_potencial, inicio_ocupado) < min(fim_potencial, fim_ocupado):
                return True  # Há conflito
        return False

    # 6. Gerar horários disponíveis
    horarios_disponiveis = []
    agora = timezone.now()

    # Itera sobre cada bloco de trabalho (ex: 09:00-12:00, 14:00-18:00)
    for bloco in blocos_de_trabalho:
        horario_atual = datetime.combine(data, bloco.hora_inicio)
        horario_fim_bloco = datetime.combine(data, bloco.hora_fim)

        # Itera dentro do bloco (ex: 09:00, 09:30, 10:00...)
        while horario_atual < horario_fim_bloco:
            inicio_potencial = timezone.make_aware(horario_atual)

            # Não mostra horários que já passaram
            if data == agora.date() and inicio_potencial < agora:
                horario_atual += timedelta(minutes=intervalo_minutos)
                continue

            fim_potencial = inicio_potencial + duracao_novo_servico

            # Verifica se o slot cabe DENTRO do bloco de trabalho
            # E se não tem conflito com agendamentos existentes
            if fim_potencial.time() <= bloco.hora_fim and not verificar_conflito(inicio_potencial, fim_potencial):
                horarios_disponiveis.append(inicio_potencial.strftime('%H:%M'))

            horario_atual += timedelta(minutes=intervalo_minutos)

    # --- FIM DA NOVA LÓGICA ---

    return JsonResponse(horarios_disponiveis, safe=False)


def dias_disponiveis(request, empreendedor_slug):
    # --- LÓGICA MODIFICADA ---
    mes_str = request.GET.get('mes')
    ano_str = request.GET.get('ano')
    servico_id = request.GET.get('servico_id')
    empreendedor_id = request.GET.get('empreendedor_id')
    # --- NOVA ADIÇÃO ---
    duracao_selecionada = request.GET.get('duracao')
    tier_id = request.GET.get('tier_id')  # ID do tier de manutenção
    # --- FIM DA ADIÇÃO ---

    if not mes_str or not ano_str or not servico_id or not empreendedor_id or not duracao_selecionada:
        return JsonResponse({'status': 'error', 'message': 'Mês, ano, serviço, profissional e duração são obrigatórios.'}, status=400)

    try:
        mes = int(mes_str)
        ano = int(ano_str)
        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)
        servico = Servico.objects.get(id=servico_id, negocio=negocio)
        profissional = EmpreendedorProfile.objects.get(
            id=empreendedor_id, negocio=negocio)

        if not servico.profissionais_que_executam.filter(id=profissional.id).exists():
            return JsonResponse({'status': 'error', 'message': 'Profissional não executa este serviço.'}, status=400)

        # 1. A duração correta (do tier ou principal) é definida AQUI
        duracao_novo_servico = timedelta(minutes=int(duracao_selecionada))
        tier_selecionado = None
        if tier_id:
            tier_selecionado = PrecoManutencao.objects.get(id=tier_id)
    except (ValueError, Servico.DoesNotExist, EmpreendedorProfile.DoesNotExist, PrecoManutencao.DoesNotExist):
        return JsonResponse({'status': 'error', 'message': 'Parâmetros inválidos.'}, status=400)

    # --- LÓGICA DE VALIDAÇÃO DE MANUTENÇÃO (REQUISIÇÃO DO USUÁRIO) ---
    cliente = None
    ultimo_agendamento_categoria = None
    if request.user.is_authenticated:
        try:
            cliente = Cliente.objects.get(user=request.user, negocio=negocio)
            # Pega o último agendamento CONCLUÍDO da MESMA categoria
            ultimo_agendamento_categoria = Agendamento.objects.filter(
                cliente=cliente,
                servico__categoria=servico.categoria,
                status__in=['Concluído', 'Confirmado']
            ).order_by('-data', '-horario').first()
        except Cliente.DoesNotExist:
            pass  # Não é cliente

    # --- INÍCIO DA NOVA LÓGICA DE HORÁRIOS ---

    dias_com_horarios = []
    num_dias = calendar.monthrange(ano, mes)[1]
    hoje = timezone.now().date()

    # --- ADICIONE ESTA LINHA ---
    # Define a data máxima que pode ser agendada
    data_limite = hoje + timedelta(days=negocio.dias_antecedencia_maxima)

    intervalo_minutos = 30  # O mesmo intervalo da outra função
    # duracao_novo_servico = timedelta(minutes=servico.duracao_minutos)

    # 1. Busca todos os blocos de trabalho do profissional
    blocos_de_trabalho_prof = HorarioTrabalho.objects.filter(
        empreendedor=profissional)

    # 2. Busca todos os agendamentos do profissional no mês
    agendamentos_prof_mes = Agendamento.objects.filter(
        empreendedor_executor=profissional,
        data__year=ano,
        data__month=mes
    )

    # Organiza em dicionários para acesso rápido
    mapa_horarios = {h.dia_da_semana: [] for h in blocos_de_trabalho_prof}
    for h in blocos_de_trabalho_prof:
        mapa_horarios[h.dia_da_semana].append(h)

    mapa_agendamentos = {d: [] for d in range(1, num_dias + 1)}
    for a in agendamentos_prof_mes:
        mapa_agendamentos[a.data.day].append(a)

    # 3. Busca todos os dias bloqueados do profissional
    dias_bloqueados_set = set(
        DiaBloqueado.objects.filter(
            empreendedor=profissional, data__year=ano, data__month=mes
        ).values_list('data', flat=True)
    )

    # --- LÓGICA DE VERIFICAÇÃO DE CONFLITO ---
    def verificar_conflito_dia(inicio_potencial, fim_potencial, agendamentos_do_dia):
        for ag in agendamentos_do_dia:
            inicio_ocupado_naive = datetime.combine(ag.data, ag.horario)
            fim_ocupado_naive = inicio_ocupado_naive + \
                timedelta(minutes=ag.duracao_final)  # <-- USA duracao_final

            if max(inicio_potencial, inicio_ocupado_naive) < min(fim_potencial, fim_ocupado_naive):
                return True
        return False
    # --- FIM DA MODIFICAÇÃO ---

    # Itera por cada dia do mês
    for dia in range(1, num_dias + 1):
        data_atual = datetime(ano, mes, dia).date()
        dia_da_semana = data_atual.weekday()

        # --- MODIFIQUE ESTE 'if' ---
        if (data_atual < hoje or
            data_atual > data_limite or
            dia_da_semana not in mapa_horarios or
                data_atual in dias_bloqueados_set):  # <-- ADICIONE ESTA VERIFICAÇÃO
            continue

        # =================================================================
        # NOVA VALIDAÇÃO (REQUISIÇÃO DO USUÁRIO)
        # "não adianta a cliente marcar o serviço de manutenção de 5 dias para uma data daqui a 15 dias"
        # =================================================================
        if tier_selecionado and ultimo_agendamento_categoria:
            # Calcula quantos dias se PASSARAM desde o último serviço ATÉ A DATA QUE ELA QUER AGENDAR
            dias_totais_desde_servico = (
                data_atual - ultimo_agendamento_categoria.data).days

            # Se os dias totais estiverem FORA do range do tier selecionado,
            # este dia é INVÁLIDO para este tier.
            if not (tier_selecionado.dias_min <= dias_totais_desde_servico <= tier_selecionado.dias_max):
                continue  # Pula este dia, ele não é válido para esta manutenção
        # --- FIM DA NOVA VALIDAÇÃO ---

        agendamentos_do_dia = mapa_agendamentos.get(dia, [])
        tem_horario_vago = False

        # Itera sobre os blocos de trabalho daquele dia (ex: manhã, tarde)
        for bloco in mapa_horarios[dia_da_semana]:
            if tem_horario_vago:  # Se já achamos um, podemos pular este bloco
                break

            horario_atual = datetime.combine(data_atual, bloco.hora_inicio)
            horario_fim_bloco = datetime.combine(data_atual, bloco.hora_fim)

            # Itera dentro do bloco (ex: 09:00, 09:30, 10:00...)
            while horario_atual < horario_fim_bloco:
                fim_potencial = horario_atual + duracao_novo_servico

                if fim_potencial.time() <= bloco.hora_fim and not verificar_conflito_dia(horario_atual, fim_potencial, agendamentos_do_dia):
                    tem_horario_vago = True
                    break  # Achamos um horário vago, podemos parar de procurar neste bloco

                horario_atual += timedelta(minutes=intervalo_minutos)

        if tem_horario_vago:
            dias_com_horarios.append(data_atual.strftime('%Y-%m-%d'))

    # --- FIM DA NOVA LÓGICA ---

    return JsonResponse(dias_com_horarios, safe=False)


# ---
# Views do Dashboard (Admin do Empreendedor)
# ---

def is_admin(user):
    return user.is_authenticated and user.is_staff


def scoped_admin_login(request, empreendedor_slug):
    """
    Realiza o login de um administrador (empreendedor/staff)
    mas APENAS se ele pertencer ao Negócio especificado no 'empreendedor_slug'.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)

    # 1. Encontra o Negócio que o usuário está TENTANDO acessar
    try:
        negocio_alvo = Negocio.objects.get(slug=empreendedor_slug)
    except Negocio.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Negócio não encontrado.'}, status=404)

    data = json.loads(request.body)
    email = data.get('email')
    password = data.get('password')

    # 2. Autentica o usuário (globalmente)
    try:
        user = User.objects.get(email=email)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Credenciais inválidas ou sem permissão de acesso.'}, status=401)

    if user.check_password(password) and user.is_staff:
        # 3. VERIFICAÇÃO CRUCIAL
        try:
            # Verifica se o usuário tem um perfil
            perfil_usuario = user.empreendedor_profile

            # Verifica se o negócio do perfil é o MESMO do slug da URL
            if perfil_usuario.negocio == negocio_alvo:
                # SUCESSO! Ele pertence a este negócio.
                login(request, user)
                return JsonResponse({
                    'status': 'success',
                    'message': 'Login bem-sucedido!',
                })
            else:
                # Ele é um admin, mas de OUTRO negócio.
                # 403 Forbidden
                return JsonResponse({'status': 'error', 'message': 'Você não tem permissão para administrar este negócio.'}, status=403)

        except EmpreendedorProfile.DoesNotExist:
            # É staff (como um superadmin) mas não tem perfil de empreendedor
            # Opcional: permitir que o superadmin logue em qualquer lugar
            if user.is_superuser:
                login(request, user)
                return JsonResponse({
                    'status': 'success',
                    'message': 'Login de Superusuário bem-sucedido!',
                })
            # Se não for superadmin, ele é apenas um staff sem perfil.
            return JsonResponse({'status': 'error', 'message': 'Este usuário não possui um perfil de empreendedor.'}, status=401)
    else:
        # Senha errada ou não é staff
        return JsonResponse({'status': 'error', 'message': 'Credenciais inválidas ou sem permissão de acesso.'}, status=401)


def global_admin_login(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        email = data.get('email')
        password = data.get('password')
        try:
            user = User.objects.get(email=email)
        except User.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Credenciais inválidas ou sem permissão de acesso.'}, status=401)

        if user.check_password(password) and user.is_staff:
            try:
                # --- CORREÇÃO AQUI ---
                _ = user.empreendedor_profile
                login(request, user)
                return JsonResponse({
                    'status': 'success',
                    'message': 'Login de administrador bem-sucedido!',
                    'user': {
                        'name': user.first_name,
                        'email': user.email
                    }
                })
            # --- CORREÇÃO AQUI ---
            except EmpreendedorProfile.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'Este usuário não possui um perfil de empreendedor associado.'}, status=401)
        else:
            return JsonResponse({'status': 'error', 'message': 'Credenciais inválidas ou sem permissão de acesso.'}, status=401)
    return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)


@user_passes_test(is_admin)
def admin_dashboard(request):
    try:
        # --- CORREÇÃO AQUI ---
        perfil = request.user.empreendedor_profile
        negocio = perfil.negocio
    except EmpreendedorProfile.DoesNotExist:
        return render(request, 'agendamentos/dashboard/index.html', {'error': 'Perfil de empreendedor não encontrado.'})

    hoje = timezone.now().date()
    base_agendamentos = Agendamento.objects.filter(servico__negocio=negocio)
    base_despesas = Despesa.objects.filter(negocio=negocio)
    agendamentos_hoje = base_agendamentos.filter(data=hoje).count()
    agendamentos_pendentes = base_agendamentos.filter(
        status='Confirmado', data__gte=hoje).count()
    pagamentos_pendentes = base_agendamentos.filter(
        status_pagamento='Pendente').count()
    inicio_mes = hoje.replace(day=1)
    ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
    fim_mes = hoje.replace(day=ultimo_dia)
    faturamento_mes = base_agendamentos.filter(
        data__range=[inicio_mes, fim_mes],
        status_pagamento='Pago'
    ).aggregate(total=Sum('servico__preco'))['total'] or 0
    despesas_mes = base_despesas.filter(
        data__range=[inicio_mes, fim_mes]
    ).aggregate(total=Sum('valor'))['total'] or 0
    context = {
        'agendamentos_hoje': agendamentos_hoje,
        'agendamentos_pendentes': agendamentos_pendentes,
        'pagamentos_pendentes': pagamentos_pendentes,
        'faturamento_mes': faturamento_mes,
        'despesas_mes': despesas_mes,
        'lucro_mes': faturamento_mes - despesas_mes,
        'hoje': hoje,
        'semana_passada': hoje - timedelta(days=7),
        'empreendedor_slug': negocio.slug
    }
    return render(request, 'agendamentos/dashboard/index.html', context)


@user_passes_test(is_admin)
def admin_calendario(request):
    try:
        # --- CORREÇÃO AQUI ---
        slug = request.user.empreendedor_profile.negocio.slug
        context = {'empreendedor_slug': slug}
        return render(request, 'agendamentos/dashboard/calendario.html', context)
    except EmpreendedorProfile.DoesNotExist:
        return render(request, 'agendamentos/dashboard/calendario.html', {'error': 'Perfil não encontrado.'})


@user_passes_test(is_admin)
def admin_financeiro(request):
    try:
        # --- CORREÇÃO AQUI ---
        slug = request.user.empreendedor_profile.negocio.slug
        context = {'empreendedor_slug': slug}
        return render(request, 'agendamentos/dashboard/financeiro.html', context)
    except EmpreendedorProfile.DoesNotExist:
        return render(request, 'agendamentos/dashboard/financeiro.html', {'error': 'Perfil não encontrado.'})


@user_passes_test(is_admin)
def admin_relatorios(request):
    try:
        # --- CORREÇÃO AQUI ---
        slug = request.user.empreendedor_profile.negocio.slug
        context = {'empreendedor_slug': slug}
        return render(request, 'agendamentos/dashboard/relatorios.html', context)
    except EmpreendedorProfile.DoesNotExist:
        return render(request, 'agendamentos/dashboard/relatorios.html', {'error': 'Perfil não encontrado.'})


# ---
# APIs do Dashboard (requerem login de admin)
# ---

@user_passes_test(is_admin)
def api_agendamentos_calendario(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse([], safe=False)

    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    start = datetime.strptime(
        start_date[:10], '%Y-%m-%d').date() if start_date else timezone.now().date()
    end = datetime.strptime(
        end_date[:10], '%Y-%m-%d').date() if end_date else (start + timedelta(days=30))

    agendamentos = Agendamento.objects.filter(
        data__range=[start, end],
        servico__negocio=negocio
    ).select_related(
        'cliente__user',
        'servico',
        'tier_manutencao'  # Garante que estamos buscando o tier
    )

    eventos = []
    for agendamento in agendamentos:
        # --- LÓGICA DE COR (Correta da última etapa) ---
        cor = '#FF9500'  # Laranja (Padrão para Pendente)

        if agendamento.status == 'Confirmado':
            # cor = '#5CCFAC' # Verde (Antigo)
            cor = '#0D99FF'  # Azul (NOVO)
        elif agendamento.status == 'Concluído':
            # cor = '#0D99FF' # Azul (Antigo)
            cor = '#5CCFAC'  # Verde (NOVO)
        elif agendamento.status == 'Cancelado':
            cor = '#FF5A5F'  # Vermelho

        # --- INÍCIO DAS CORREÇÕES ---

        # 1. Define o NOME e o ID corretos
        nome_servico = ""
        servico_tier_id_str = ""
        if agendamento.tier_manutencao:
            # Lógica IDÊNTICA a 'api_admin_get_form_data'
            nome_servico = f"{agendamento.servico.nome} - {agendamento.tier_manutencao.nome_tier}"
            servico_tier_id_str = f"tier_{agendamento.tier_manutencao.id}"
        else:
            nome_servico = agendamento.servico.nome
            servico_tier_id_str = f"service_{agendamento.servico.id}"

        # 2. Define a DURAÇÃO correta (do campo 'duracao_final' salvo no .save())
        # Fallback de 15 min se algo der errado
        duracao_real = agendamento.duracao_final or 15

        # 3. Define o PREÇO correto (do campo 'preco_final' salvo no .save())
        preco_real = agendamento.preco_final or 0.0

        # --- FIM DAS CORREÇÕES ---

        eventos.append({
            'id': agendamento.id,
            # <-- CORRIGIDO
            'title': f"{agendamento.cliente.user.get_full_name()} - {nome_servico}",
            'start': f"{agendamento.data.isoformat()}T{agendamento.horario.isoformat()}",
            'end': (datetime.combine(agendamento.data, agendamento.horario) +
                    # <-- CORRIGIDO
                    timedelta(minutes=duracao_real)).isoformat(),
            'color': cor,
            'extendedProps': {
                'cliente': agendamento.cliente.user.get_full_name(),
                'email': agendamento.cliente.user.email,
                'telefone': agendamento.cliente.telefone,
                'servico': nome_servico,  # <-- CORRIGIDO (nome para exibição)
                # <-- NOVO (para pré-seleção)
                'servico_tier_id': servico_tier_id_str,
                'preco': float(preco_real),  # <-- CORRIGIDO
                'status': agendamento.status,
                'status_pagamento': agendamento.status_pagamento,
                'observacoes': agendamento.observacoes or ''
            }
        })
    return JsonResponse(eventos, safe=False)


@user_passes_test(is_admin)
def api_resumo_financeiro(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    periodo = request.GET.get('periodo', 'mes')
    hoje = timezone.now().date()
    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            inicio = hoje.replace(day=1)
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            fim = hoje.replace(day=ultimo_dia)

    base_agendamentos = Agendamento.objects.filter(servico__negocio=negocio)
    base_despesas = Despesa.objects.filter(negocio=negocio)

    # CORREÇÃO: Usar 'preco_final' para somar o valor real do agendamento (ex: valor da manutenção)
    # em vez do preço base do serviço 'servico__preco'.
    faturamento = base_agendamentos.filter(
        data__range=[inicio, fim],
        status_pagamento='Pago'
    ).aggregate(total=Sum('preco_final'))['total'] or 0

    faturamento_pendente = base_agendamentos.filter(
        data__range=[inicio, fim],
        status_pagamento='Pendente',
        status__in=['Confirmado', 'Concluído']
    ).aggregate(total=Sum('preco_final'))['total'] or 0

    despesas = base_despesas.filter(
        data__range=[inicio, fim]
    ).aggregate(total=Sum('valor'))['total'] or 0

    total_atendimentos = base_agendamentos.filter(
        data__range=[inicio, fim],
        status__in=['Confirmado', 'Concluído']
    ).count()

    atendimentos_concluidos = base_agendamentos.filter(
        data__range=[inicio, fim],
        status='Concluído'
    ).count()

    # CORREÇÃO PRINCIPAL: Calcular o valor total somando o preço final dos agendamentos
    servicos_populares_query = base_agendamentos.filter(
        data__range=[inicio, fim],
        status__in=['Confirmado', 'Concluído']
    ).values('servico__nome').annotate(
        qtd=Count('id'),
        # Soma o valor real arrecadado por esse serviço
        valor_total=Sum('preco_final')
    ).order_by('-valor_total')[:5]

    # Formata a lista para o JSON
    servicos_populares_list = [
        {
            'servico__nome': item['servico__nome'],
            'total': item['qtd'],
            # Envia o valor monetário já calculado pelo banco
            'valor_monetario': float(item['valor_total'] or 0)
        }
        for item in servicos_populares_query
    ]

    return JsonResponse({
        'periodo': {
            'inicio': inicio.isoformat(),
            'fim': fim.isoformat()
        },
        'financeiro': {
            'faturamento': float(faturamento),
            'faturamento_pendente': float(faturamento_pendente),
            'despesas': float(despesas),
            'lucro': float(faturamento - despesas)
        },
        'atendimentos': {
            'total': total_atendimentos,
            'concluidos': atendimentos_concluidos,
            'pagos': base_agendamentos.filter(
                data__range=[inicio, fim],
                status_pagamento='Pago'
            ).count()
        },
        'servicos_populares': servicos_populares_list
    })


@user_passes_test(is_admin)
def api_faturamento(request):
    try:
        # --- CORREÇÃO AQUI ---
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    periodo = request.GET.get('periodo', 'mes')
    tipo = request.GET.get('tipo', 'diario')
    hoje = timezone.now().date()
    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            inicio = hoje.replace(day=1)
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            fim = hoje.replace(day=ultimo_dia)

    base_agendamentos = Agendamento.objects.filter(servico__negocio=negocio)

    if tipo == 'diario':
        agendamentos = base_agendamentos.filter(
            data__range=[inicio, fim],
            status_pagamento='Pago'
        ).values('data').annotate(
            total=Sum('servico__preco'),
            quantidade=Count('id')
        ).order_by('data')
        dados = [
            {
                'data': item['data'].isoformat(),
                'total': float(item['total']),
                'quantidade': item['quantidade']
            }
            for item in agendamentos
        ]
    elif tipo == 'mensal':
        dados = []
        for mes in range(1, 13):
            if mes < inicio.month or mes > fim.month:
                continue
            inicio_mes = datetime(inicio.year, mes, 1).date()
            fim_mes = datetime(inicio.year, mes, calendar.monthrange(
                inicio.year, mes)[1]).date()
            total = base_agendamentos.filter(
                data__range=[inicio_mes, fim_mes],
                status_pagamento='Pago'
            ).aggregate(
                total=Sum('servico__preco'),
                quantidade=Count('id')
            )
            dados.append({
                'mes': mes,
                'nome_mes': calendar.month_name[mes],
                'total': float(total['total'] or 0),
                'quantidade': total['quantidade'] or 0
            })
    else:
        servicos = base_agendamentos.filter(
            data__range=[inicio, fim],
            status_pagamento='Pago'
        ).values('servico__nome').annotate(
            total=Sum('servico__preco'),
            quantidade=Count('id')
        ).order_by('-total')
        dados = [
            {
                'servico': item['servico__nome'],
                'total': float(item['total']),
                'quantidade': item['quantidade']
            }
            for item in servicos
        ]
    return JsonResponse({
        'periodo': {
            'inicio': inicio.isoformat(),
            'fim': fim.isoformat()
        },
        'tipo': tipo,
        'dados': dados
    })


@user_passes_test(is_admin)
def api_despesas(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    periodo = request.GET.get('periodo', 'mes')
    categoria = request.GET.get('categoria', None)
    hoje = timezone.now().date()

    # Definição das datas de início e fim
    inicio = hoje.replace(day=1)
    fim = hoje.replace(day=calendar.monthrange(hoje.year, hoje.month)[1])

    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        # Já definido acima como padrão
        pass
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:
        # Personalizado
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            try:
                inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
                fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
            except ValueError:
                pass

    # Passamos a data final do filtro para o processador de recorrências.
    processar_despesas_recorrentes(negocio, data_limite_solicitada=fim)

    base_despesas = Despesa.objects.filter(negocio=negocio)
    filtros = {'data__range': [inicio, fim]}
    if categoria:
        filtros['categoria'] = categoria

    # CORREÇÃO AQUI: Alterado de '-data' para 'data' (ascendente)
    # Isso exibe da data mais antiga (ou mais próxima) para a mais distante/futura
    despesas = base_despesas.filter(**filtros).order_by('data')

    resumo_categorias = base_despesas.filter(
        data__range=[inicio, fim]
    ).values('categoria').annotate(
        total=Sum('valor'),
        quantidade=Count('id')
    ).order_by('-total')

    return JsonResponse({
        'periodo': {
            'inicio': inicio.isoformat(),
            'fim': fim.isoformat()
        },
        'despesas': [
            {
                'id': despesa.id,
                'descricao': despesa.descricao,
                'valor': float(despesa.valor),
                'data': despesa.data.isoformat(),
                'categoria': despesa.categoria,
                'pago': despesa.pago,
                'comprovante': despesa.comprovante.url if despesa.comprovante else None
            }
            for despesa in despesas
        ],
        'resumo_categorias': [
            {
                'categoria': categoria['categoria'],
                'total': float(categoria['total']),
                'quantidade': categoria['quantidade']
            }
            for categoria in resumo_categorias
        ],
        'total': float(despesas.aggregate(total=Sum('valor'))['total'] or 0)
    })


@user_passes_test(is_admin)
def api_listar_recorrencias(request):
    """Lista todas as regras de recorrência ativas (que gerarão cobranças futuras)."""
    try:
        negocio = request.user.empreendedor_profile.negocio
        hoje = timezone.now().date()

        # Define o fim do mês atual
        ultimo_dia_mes = calendar.monthrange(hoje.year, hoje.month)[1]
        fim_deste_mes = hoje.replace(day=ultimo_dia_mes)

        # CORREÇÃO:
        # Filtra recorrências que não têm data fim (infinitas)
        # OU cuja data fim seja MAIOR que o fim deste mês.
        # Se a data fim for igual ao fim deste mês (finalizada), ela não aparece mais na lista de gestão.
        recorrencias = DespesaRecorrente.objects.filter(
            negocio=negocio
        ).filter(
            Q(data_fim__isnull=True) | Q(data_fim__gt=fim_deste_mes)
        ).order_by('dia_vencimento')

        data = [{
            'id': rec.id,
            'descricao': rec.descricao,
            'valor': float(rec.valor),
            'categoria': rec.categoria,
            'dia_vencimento': rec.dia_vencimento,
            'data_inicio': rec.data_inicio.isoformat(),
            'data_fim': rec.data_fim.isoformat() if rec.data_fim else None
        } for rec in recorrencias]

        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_gerenciar_recorrencia_detalhe(request, recorrencia_id):
    """Edita valor ou encerra uma recorrência e limpa lançamentos futuros."""
    try:
        negocio = request.user.empreendedor_profile.negocio
        recorrencia = get_object_or_404(
            DespesaRecorrente, id=recorrencia_id, negocio=negocio)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Recorrência não encontrada.'}, status=404)

    if request.method == 'POST':
        data = json.loads(request.body)
        acao = data.get('acao')

        if acao == 'encerrar':
            # 1. Define a data fim para o ÚLTIMO DIA DO MÊS ATUAL
            # Isso faz com que ela pare de aparecer na lista (pelo filtro da função acima)
            hoje = timezone.now().date()
            ultimo_dia_mes = calendar.monthrange(hoje.year, hoje.month)[1]
            fim_deste_mes = hoje.replace(day=ultimo_dia_mes)

            recorrencia.data_fim = fim_deste_mes
            recorrencia.save()

            # 2. LIMPEZA DE FUTURO (CRUCIAL PARA PODER USAR O MESMO NOME DEPOIS)
            # Remove as despesas individuais que JÁ tinham sido geradas no banco para datas futuras
            # Isso limpa a agenda para que você possa criar uma nova regra com o mesmo nome sem duplicatas.

            nome_gerado = f"{recorrencia.descricao} (Recorrente)"

            Despesa.objects.filter(
                negocio=negocio,
                descricao=nome_gerado,
                categoria=recorrencia.categoria,
                valor=recorrencia.valor,
                data__gt=fim_deste_mes,          # Apaga tudo que estiver DEPOIS deste mês
                pago=False                       # Segurança: não apaga se já foi marcado como pago
            ).delete()

            return JsonResponse({'status': 'success', 'message': 'Cobrança recorrente finalizada. Lançamentos futuros removidos.'})

        elif acao == 'editar':
            # Atualiza os dados para as PRÓXIMAS gerações
            if 'valor' in data:
                recorrencia.valor = data['valor']
            if 'descricao' in data:
                recorrencia.descricao = data['descricao']
            if 'dia_vencimento' in data:
                recorrencia.dia_vencimento = data['dia_vencimento']

            recorrencia.save()
            return JsonResponse({'status': 'success', 'message': 'Recorrência atualizada para os próximos meses.'})

    return JsonResponse({'status': 'error', 'message': 'Método inválido'}, status=405)


@user_passes_test(is_admin)
def api_atualizar_pagamento(request, agendamento_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)
    try:
        # --- CORREÇÃO AQUI ---
        negocio = request.user.empreendedor_profile.negocio
        agendamento = get_object_or_404(
            Agendamento, id=agendamento_id, servico__negocio=negocio)
        dados = json.loads(request.body)
        if 'status_pagamento' in dados:
            agendamento.status_pagamento = dados['status_pagamento']
        if 'status' in dados:
            agendamento.status = dados['status']
        if 'observacoes' in dados:
            agendamento.observacoes = dados['observacoes']
        agendamento.save()
        return JsonResponse({
            'status': 'success',
            'message': 'Agendamento atualizado com sucesso',
            'agendamento': {
                'id': agendamento.id,
                'status': agendamento.status,
                'status_pagamento': agendamento.status_pagamento
            }
        })
    # --- CORREÇÃO AQUI ---
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Agendamento.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Agendamento não encontrado ou não pertence a você.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_registrar_despesa(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)
    try:
        negocio = request.user.empreendedor_profile.negocio
        dados = json.loads(request.body)

        data_obj = datetime.strptime(dados['data'], '%Y-%m-%d').date()

        # 1. Cria a despesa atual (sempre cria a primeira)
        nova_despesa = Despesa(
            negocio=negocio,
            descricao=dados['descricao'],
            valor=dados['valor'],
            data=data_obj,
            categoria=dados['categoria'],
            pago=dados.get('pago', False)
        )
        nova_despesa.save()

        # 2. Verifica se é recorrente
        is_recorrente = dados.get('recorrente', False)
        if is_recorrente:
            data_fim = None
            if dados.get('data_fim_recorrencia'):
                data_fim = datetime.strptime(
                    dados['data_fim_recorrencia'], '%Y-%m-%d').date()

            DespesaRecorrente.objects.create(
                negocio=negocio,
                descricao=dados['descricao'],
                valor=dados['valor'],
                categoria=dados['categoria'],
                data_inicio=data_obj,
                dia_vencimento=data_obj.day,
                data_fim=data_fim,
                # Marca que a deste mês já foi gerada (a 'nova_despesa' acima)
                ultima_geracao=data_obj
            )

        return JsonResponse({
            'status': 'success',
            'message': 'Despesa registrada com sucesso',
            'despesa': {
                'id': nova_despesa.id,
                'descricao': nova_despesa.descricao,
                'valor': float(nova_despesa.valor),
                'data': nova_despesa.data.isoformat(),
                'categoria': nova_despesa.categoria,
                'pago': nova_despesa.pago
            }
        })
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_atualizar_despesa(request, despesa_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)
    try:
        negocio = request.user.empreendedor_profile.negocio
        despesa = get_object_or_404(Despesa, id=despesa_id, negocio=negocio)

        dados = json.loads(request.body)

        if 'descricao' in dados:
            despesa.descricao = dados['descricao']
        if 'valor' in dados:
            despesa.valor = dados['valor']
        if 'data' in dados:
            despesa.data = datetime.strptime(dados['data'], '%Y-%m-%d').date()
        if 'categoria' in dados:
            despesa.categoria = dados['categoria']
        if 'pago' in dados:
            despesa.pago = dados['pago']

        despesa.save()

        # Nota: A edição de uma despesa individual NÃO afeta a regra de recorrência
        # (DespesaRecorrente) para manter a integridade do histórico.

        return JsonResponse({
            'status': 'success',
            'message': 'Despesa atualizada com sucesso',
            'despesa': {
                'id': despesa.id,
                'descricao': despesa.descricao,
                'valor': float(despesa.valor),
                'data': despesa.data.isoformat(),
                'categoria': despesa.categoria,
                'pago': despesa.pago
            }
        })
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Despesa.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Despesa não encontrada.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_deletar_despesa(request, despesa_id):
    """
    Exclui uma despesa específica.
    """
    if request.method != 'DELETE':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)

    try:
        # Garante que o admin logado só possa deletar despesas do seu negócio
        negocio = request.user.empreendedor_profile.negocio
        despesa = get_object_or_404(
            Despesa, id=despesa_id, negocio=negocio)

        # Exclui o objeto do banco de dados
        despesa.delete()

        return JsonResponse({
            'status': 'success',
            'message': 'Despesa excluída com sucesso'
        })

    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Despesa.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Despesa não encontrada ou não pertence a você.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_despesa(request, despesa_id):
    try:
        # --- CORREÇÃO AQUI ---
        negocio = request.user.empreendedor_profile.negocio
        despesa = get_object_or_404(
            Despesa, id=despesa_id, negocio=negocio)
        return JsonResponse({
            'id': despesa.id,
            'descricao': despesa.descricao,
            'valor': float(despesa.valor),
            'data': despesa.data.isoformat(),
            'categoria': despesa.categoria,
            'pago': despesa.pago,
            'comprovante': despesa.comprovante.url if despesa.comprovante else None
        })
    # --- CORREÇÃO AQUI ---
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Despesa.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Despesa não encontrada ou não pertence a você.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_agendamentos_pagamento(request):
    try:
        # --- CORREÇÃO AQUI ---
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    periodo = request.GET.get('periodo', 'mes')
    hoje = timezone.now().date()
    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            inicio = hoje.replace(day=1)
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            fim = hoje.replace(day=ultimo_dia)

    base_agendamentos = Agendamento.objects.filter(servico__negocio=negocio)

    pendentes = base_agendamentos.filter(
        data__range=[inicio, fim],
        status_pagamento='Pendente',
        status__in=['Confirmado', 'Concluído']
    ).select_related('cliente__user', 'servico')
    pagos = base_agendamentos.filter(
        data__range=[inicio, fim],
        status_pagamento='Pago'
    ).select_related('cliente__user', 'servico')

    return JsonResponse({
        'periodo': {
            'inicio': inicio.isoformat(),
            'fim': fim.isoformat()
        },
        'pendentes': [
            {
                'id': agendamento.id,
                'cliente': f"{agendamento.cliente.user.first_name} {agendamento.cliente.user.last_name}",
                'servico': agendamento.servico.nome,
                'data': agendamento.data.isoformat(),
                'horario': agendamento.horario.strftime('%H:%M'),
                'valor': float(agendamento.servico.preco),
                'status': agendamento.status,
                'status_pagamento': agendamento.status_pagamento
            }
            for agendamento in pendentes
        ],
        'pagos': [
            {
                'id': agendamento.id,
                'cliente': f"{agendamento.cliente.user.first_name} {agendamento.cliente.user.last_name}",
                'servico': agendamento.servico.nome,
                'data': agendamento.data.isoformat(),
                'horario': agendamento.horario.strftime('%H:%M'),
                'valor': float(agendamento.servico.preco),
                'status': agendamento.status,
                'status_pagamento': agendamento.status_pagamento
            }
            for agendamento in pagos
        ]
    })

# --- View de Lista de Agendamentos (Exemplo) ---
# Esta view é um exemplo de como você pode listar agendamentos.
# Ela não está sendo usada pelo seu SPA principal, mas é útil para depuração.


def lista_agendamentos(request, empreendedor_slug):
    negocio = get_object_or_404(Negocio, slug=empreendedor_slug)

    # Filtra agendamentos por negócio
    agendamentos = Agendamento.objects.filter(servico__negocio=negocio).select_related(
        'cliente__user', 'servico'
    ).order_by('-data', '-horario')

    data = [
        {
            'id': agendamento.id,
            'client': f"{agendamento.cliente.user.first_name} {agendamento.cliente.user.last_name}",
            'service': agendamento.servico.nome,
            'date': agendamento.data.strftime('%Y-%m-%d'),
            'time': agendamento.horario.strftime('%H:%M'),
            'status': agendamento.status
        } for agendamento in agendamentos
    ]
    return JsonResponse(data, safe=False)


# 1. VIEW DA PÁGINA DE GESTÃO (RENDERIZA O HTML)
# ---
@user_passes_test(is_admin)
def admin_gestao(request):
    try:
        # Passa o slug para o template (para o link "Voltar ao Site")
        slug = request.user.empreendedor_profile.negocio.slug
        context = {'empreendedor_slug': slug}
        return render(request, 'agendamentos/dashboard/gestao.html', context)
    except EmpreendedorProfile.DoesNotExist:
        return render(request, 'agendamentos/dashboard/gestao.html', {'error': 'Perfil não encontrado.'})

# ---
# 2. API PARA GERENCIAR A LISTA DE SERVIÇOS (LER E CRIAR)
# ---


@user_passes_test(is_admin)
def api_gestao_servicos(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    # --- LER (GET) ---
    if request.method == 'GET':
        servicos = Servico.objects.filter(
            negocio=negocio).prefetch_related('profissionais_que_executam')
        data = [
            {
                'id': servico.id,
                'nome': servico.nome,
                'preco': float(servico.preco),
                'duracao_minutos': servico.duracao_minutos,
                'descricao': servico.descricao,
                'imagem_url': servico.imagem.url if servico.imagem else None,
                'profissionais_ids': [p.id for p in servico.profissionais_que_executam.all()],
                # --- NOVA ADIÇÃO ---
                'categoria_id': servico.categoria_id,
                'categoria_nome': servico.categoria.nome if servico.categoria else None
                # --- FIM DA ADIÇÃO ---
            } for servico in servicos
        ]
        return JsonResponse(data, safe=False)

    # --- MODIFIQUE O 'POST' para incluir a categoria ---
    if request.method == 'POST':
        # MUDANÇA: Lendo de request.POST (Form Data) em vez de JSON
        data = request.POST
        try:
            # --- NOVA ADIÇÃO ---
            categoria_id = data.get('categoria_id')
            categoria = None
            if categoria_id:
                categoria = Categoria.objects.get(
                    id=categoria_id, negocio=negocio)
            # --- FIM DA ADIÇÃO ---

            novo_servico = Servico.objects.create(
                negocio=negocio,
                nome=data['nome'],
                descricao=data.get('descricao', ''),
                preco=data['preco'],
                duracao_minutos=data['duracao_minutos'],
                percentual_adiantamento=data.get('percentual_adiantamento', 0),
                categoria=categoria  # <-- ADICIONADO
            )

            # ADICIONADO: Verifica se um arquivo de imagem foi enviado
            if 'imagem' in request.FILES:
                novo_servico.imagem = request.FILES['imagem']

            novo_servico.save()  # Salva a imagem

            # Associa os profissionais (lendo a string do form data)
            profissionais_ids_str = data.get('profissionais_ids', '')
            if profissionais_ids_str:
                profissionais_ids = [int(id)
                                     for id in profissionais_ids_str.split(',')]
                profissionais = EmpreendedorProfile.objects.filter(
                    negocio=negocio, id__in=profissionais_ids)
                novo_servico.profissionais_que_executam.set(profissionais)

            return JsonResponse({'status': 'success', 'message': 'Serviço criado com sucesso.'}, status=201)
        except Categoria.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Categoria não encontrada.'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ---
# 3. API PARA GERENCIAR UM SERVIÇO ESPECÍFICO (EDITAR, EXCLUIR, LER DETALHE)
# ---


@user_passes_test(is_admin)
def api_gestao_servico_detalhe(request, servico_id):
    try:
        negocio = request.user.empreendedor_profile.negocio
        servico = get_object_or_404(Servico, id=servico_id, negocio=negocio)
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Servico.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Serviço não encontrado.'}, status=404)

    # --- LER DETALHE (GET) ---
    if request.method == 'GET':
        data = {
            'id': servico.id,
            'nome': servico.nome,
            'preco': float(servico.preco),
            'duracao_minutos': servico.duracao_minutos,
            'descricao': servico.descricao,
            'imagem_url': servico.imagem.url if servico.imagem else None,
            'profissionais_ids': [p.id for p in servico.profissionais_que_executam.all()],
            'percentual_adiantamento': servico.percentual_adiantamento,
            # --- NOVA ADIÇÃO ---
            'categoria_id': servico.categoria_id
            # --- FIM DA ADIÇÃO ---
        }
        return JsonResponse(data)

    # --- EDITAR (POST) ---
    if request.method == 'POST':
        # MUDANÇA: Lendo de request.POST (Form Data) em vez de JSON
        data = request.POST
        try:
            servico.nome = data['nome']
            servico.descricao = data.get('descricao', '')
            servico.preco = data['preco']
            servico.duracao_minutos = data['duracao_minutos']
            servico.percentual_adiantamento = data.get(
                'percentual_adiantamento', 0)

            # ADICIONADO: Verifica se um arquivo de imagem foi enviado
            if 'imagem' in request.FILES:
                servico.imagem = request.FILES['imagem']

            categoria_id = data.get('categoria_id')
            categoria = None
            if categoria_id:
                categoria = Categoria.objects.get(
                    id=categoria_id, negocio=negocio)
            servico.categoria = categoria  # <-- ADICIONADO

            servico.save()  # Salva o texto e a nova imagem

            # Atualiza os profissionais associados
            profissionais_ids_str = data.get('profissionais_ids', '')
            if profissionais_ids_str:
                profissionais_ids = [int(id)
                                     for id in profissionais_ids_str.split(',')]
                profissionais = EmpreendedorProfile.objects.filter(
                    negocio=negocio, id__in=profissionais_ids)
                servico.profissionais_que_executam.set(profissionais)
            else:
                # Se enviado vazio, remove todos
                servico.profissionais_que_executam.clear()

            return JsonResponse({'status': 'success', 'message': 'Serviço atualizado com sucesso.'})
        except Categoria.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'Categoria não encontrada.'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # --- EXCLUIR (DELETE) ---
    if request.method == 'DELETE':
        try:
            servico.delete()
            return JsonResponse({'status': 'success', 'message': 'Serviço excluído com sucesso.'}, status=204)
        except Exception as e:
            # Proteção contra deleção de serviço com agendamento
            if 'FOREIGN KEY constraint' in str(e):
                return JsonResponse({'status': 'error', 'message': 'Não é possível excluir este serviço, pois ele já possui agendamentos.'}, status=400)
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# =================================================================
# NOVA ADIÇÃO: VIEWS DA API DE CATEGORIA
# =================================================================
@user_passes_test(is_admin)
def api_gestao_categorias(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    if request.method == 'GET':
        categorias = Categoria.objects.filter(negocio=negocio)
        data = [{'id': c.id, 'nome': c.nome} for c in categorias]
        return JsonResponse(data, safe=False)

    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            nova_cat = Categoria.objects.create(
                negocio=negocio,
                nome=data['nome']
            )
            return JsonResponse({'status': 'success', 'message': 'Categoria criada!', 'id': nova_cat.id}, status=201)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_gestao_categoria_detalhe(request, categoria_id):
    try:
        negocio = request.user.empreendedor_profile.negocio
        categoria = get_object_or_404(
            Categoria, id=categoria_id, negocio=negocio)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Categoria não encontrada.'}, status=404)

    if request.method == 'POST':  # Editar
        data = json.loads(request.body)
        categoria.nome = data.get('nome', categoria.nome)
        categoria.save()
        return JsonResponse({'status': 'success', 'message': 'Categoria atualizada.'})

    if request.method == 'DELETE':
        # A categoria está com on_delete=models.SET_NULL no Servico,
        # então apagar aqui é seguro e não deletará serviços.
        categoria.delete()
        return JsonResponse({'status': 'success', 'message': 'Categoria excluída.'}, status=204)

# =================================================================
# NOVA ADIÇÃO: VIEWS DA API DE PREÇOS DE MANUTENÇÃO
# =================================================================


@user_passes_test(is_admin)
def api_gestao_precos_manutencao(request, servico_id):
    try:
        negocio = request.user.empreendedor_profile.negocio
        servico = get_object_or_404(Servico, id=servico_id, negocio=negocio)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Serviço não encontrado.'}, status=404)

    if request.method == 'GET':
        precos = PrecoManutencao.objects.filter(servico_pai=servico)
        data = [{
            'id': p.id,
            'nome_tier': p.nome_tier,
            'dias_min': p.dias_min,
            'dias_max': p.dias_max,
            'preco': float(p.preco),
            'duracao_minutos': p.duracao_minutos
        } for p in precos]
        return JsonResponse(data, safe=False)

    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            novo_preco = PrecoManutencao(
                servico_pai=servico,
                nome_tier=data['nome_tier'],
                dias_min=data['dias_min'],
                dias_max=data['dias_max'],
                preco=data['preco'],
                duracao_minutos=data['duracao_minutos'],
                percentual_adiantamento=data.get('percentual_adiantamento', 0)
            )
            novo_preco.clean()  # Roda a validação do modelo
            novo_preco.save()
            return JsonResponse({'status': 'success', 'message': 'Preço de manutenção adicionado!'}, status=201)
        except ValidationError as e:
            return JsonResponse({'status': 'error', 'message': e.message}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@user_passes_test(is_admin)
def api_gestao_preco_manutencao_detalhe(request, preco_id):
    try:
        negocio = request.user.empreendedor_profile.negocio
        preco = get_object_or_404(
            PrecoManutencao, id=preco_id, servico_pai__negocio=negocio)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Preço não encontrado.'}, status=404)

    if request.method == 'POST':  # Editar
        data = json.loads(request.body)
        try:
            preco.nome_tier = data.get('nome_tier', preco.nome_tier)
            preco.dias_min = data.get('dias_min', preco.dias_min)
            preco.dias_max = data.get('dias_max', preco.dias_max)
            preco.preco = data.get('preco', preco.preco)
            preco.duracao_minutos = data.get(
                'duracao_minutos', preco.duracao_minutos)
            preco.percentual_adiantamento = data.get(
                'percentual_adiantamento', preco.percentual_adiantamento)
            preco.clean()
            preco.save()
            return JsonResponse({'status': 'success', 'message': 'Preço atualizado.'})
        except ValidationError as e:
            return JsonResponse({'status': 'error', 'message': e.message}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    if request.method == 'DELETE':
        preco.delete()
        return JsonResponse({'status': 'success', 'message': 'Preço excluído.'}, status=204)

# ---
# 4. API PARA LER A EQUIPE (USADO NO MODAL DE SERVIÇOS)
# ---


@user_passes_test(is_admin)
def api_gestao_equipe(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    # --- LER (GET) ---
    if request.method == 'GET':
        equipe = EmpreendedorProfile.objects.filter(
            negocio=negocio).select_related('user')
        data = [
            {
                'id': membro.id,
                'nome': membro.user.get_full_name() or membro.user.username,
                'email': membro.user.email,
                # --- ADICIONADO ---
                'foto_url': membro.foto.url if membro.foto else None
            } for membro in equipe
        ]
        return JsonResponse(data, safe=False)

    # --- CONVIDAR / CRIAR (POST) ---
    if request.method == 'POST':
        # MUDANÇA: Lendo de request.POST (Form Data)
        data = request.POST

        email = data.get('email')
        password = data.get('password')

        # --- Validações ---
        if not email or not password or not data.get('nome'):
            return JsonResponse({'status': 'error', 'message': 'Nome, email e senha inicial são obrigatórios.'}, status=400)
        if User.objects.filter(email=email).exists():
            return JsonResponse({'status': 'error', 'message': 'Já existe um usuário com este email no sistema.'}, status=400)

        try:
            # 1. Cria o User
            novo_usuario = User.objects.create_user(
                username=email,
                email=email,
                password=password,
                first_name=data.get('nome'),
                last_name=data.get('sobrenome', '')
            )
            novo_usuario.is_staff = True
            novo_usuario.save()

            # 3. Cria o Perfil de Empreendedor
            novo_perfil = EmpreendedorProfile.objects.create(
                user=novo_usuario,
                negocio=negocio,
                telefone=data.get('telefone', '')
            )

            # --- ADICIONADO: Salva a foto ---
            if 'foto' in request.FILES:
                novo_perfil.foto = request.FILES['foto']
                novo_perfil.save()
            # --- FIM DA ADIÇÃO ---

            return JsonResponse({'status': 'success', 'message': 'Novo membro da equipe adicionado com sucesso!'}, status=201)

        except Exception as e:
            # Se algo der errado (ex: username duplicado), desfaz a criação do usuário
            if 'novo_usuario' in locals() and novo_usuario:
                novo_usuario.delete()
            return JsonResponse({'status': 'error', 'message': f'Erro ao criar usuário: {str(e)}'}, status=400)


# agendamentos/views.py

@user_passes_test(is_admin)
def api_gestao_horarios(request):
    try:
        profissional = request.user.empreendedor_profile
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    if request.method == 'GET':
        horarios = HorarioTrabalho.objects.filter(
            empreendedor=profissional).order_by('dia_da_semana', 'hora_inicio')
        data = [{
            'id': h.id,
            'dia_da_semana': h.dia_da_semana,
            'dia_nome': h.get_dia_da_semana_display(),
            'hora_inicio': h.hora_inicio.strftime('%H:%M'),
            'hora_fim': h.hora_fim.strftime('%H:%M'),
        } for h in horarios]
        return JsonResponse(data, safe=False)

    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # 1. Processa os horários principais
            jornada_inicio = datetime.strptime(
                data['hora_inicio'], '%H:%M').time()
            jornada_fim = datetime.strptime(data['hora_fim'], '%H:%M').time()

            # 2. Verifica se tem almoço
            almoco_inicio = None
            almoco_fim = None
            tem_almoco = False

            if data.get('almoco_inicio') and data.get('almoco_fim'):
                almoco_inicio = datetime.strptime(
                    data['almoco_inicio'], '%H:%M').time()
                almoco_fim = datetime.strptime(
                    data['almoco_fim'], '%H:%M').time()

                # Validações do Almoço
                if almoco_inicio >= almoco_fim:
                    return JsonResponse({'status': 'error', 'message': 'O fim do almoço deve ser após o início.'}, status=400)
                if almoco_inicio <= jornada_inicio or almoco_fim >= jornada_fim:
                    return JsonResponse({'status': 'error', 'message': 'O horário de almoço deve estar DENTRO da jornada de trabalho.'}, status=400)
                tem_almoco = True

            # 3. Define quais blocos serão criados
            blocos_para_criar = []

            if tem_almoco:
                # Cria dois blocos: Manhã e Tarde
                blocos_para_criar.append(
                    (jornada_inicio, almoco_inicio))  # Bloco 1 (Ex: 08-12)
                # Bloco 2 (Ex: 13-17)
                blocos_para_criar.append((almoco_fim, jornada_fim))
            else:
                # Cria um bloco só
                blocos_para_criar.append((jornada_inicio, jornada_fim))

            # 4. Salva os blocos (com validação de conflito)
            count_criados = 0

            with transaction.atomic():  # Garante que ou salva tudo ou nada
                for inicio, fim in blocos_para_criar:
                    # Validação de conflito (igual fizemos antes)
                    conflitos = HorarioTrabalho.objects.filter(
                        empreendedor=profissional,
                        dia_da_semana=data['dia_da_semana'],
                        hora_inicio__lt=fim,
                        hora_fim__gt=inicio
                    )

                    if conflitos.exists():
                        # Se der conflito em um dos blocos, aborta tudo
                        raise ValidationError(
                            f"Conflito de horário! O período {inicio.strftime('%H:%M')}-{fim.strftime('%H:%M')} bate com outro já existente.")

                    HorarioTrabalho.objects.create(
                        empreendedor=profissional,
                        dia_da_semana=data['dia_da_semana'],
                        hora_inicio=inicio,
                        hora_fim=fim
                    )
                    count_criados += 1

            msg = "Horário criado!" if count_criados == 1 else "Jornada criada com intervalo de almoço!"
            return JsonResponse({'status': 'success', 'message': msg}, status=201)

        except ValidationError as e:
            return JsonResponse({'status': 'error', 'message': e.message}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Erro ao salvar: {str(e)}'}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)


@user_passes_test(is_admin)
def api_gestao_horario_detalhe(request, horario_id):
    """
    API para EXCLUIR (ou futuramente editar) um horário específico.
    """
    try:
        profissional = request.user.empreendedor_profile
        horario = get_object_or_404(
            HorarioTrabalho, id=horario_id, empreendedor=profissional)
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except HorarioTrabalho.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Horário não encontrado.'}, status=404)

    # --- EXCLUIR (DELETE) ---
    if request.method == 'DELETE':
        try:
            horario.delete()
            return JsonResponse({'status': 'success', 'message': 'Horário removido.'}, status=204)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)


def hex_to_hsl_string(hex_color):
    """Converte #RRGGBB para uma string HSL 'H, S%, L%'"""
    hex_color = hex_color.lstrip('#')
    try:
        r, g, b = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
        r, g, b = [x / 255.0 for x in (r, g, b)]
        # O colorsys usa HLS, que é um pouco diferente de HSL
        h, l, s = colorsys.rgb_to_hls(r, g, b)
        # Converte para o formato HSL do CSS
        return f"{int(h * 360)}, {int(s * 100)}%, {int(l * 100)}%"
    except Exception:
        # Retorna o padrão (verde-menta) em caso de erro
        return "160, 41%, 58%"


def api_negocio_info(request, empreendedor_slug):
    """
    Retorna as informações públicas do negócio, incluindo as cores.
    """
    negocio = get_object_or_404(Negocio, slug=empreendedor_slug)

    # --- LÓGICA DA COR PRIMÁRIA (JÁ EXISTE) ---
    cor_primaria_hex = negocio.cor_primaria
    cor_primaria_hsl = hex_to_hsl_string(cor_primaria_hex)

    # --- ADICIONE ESTAS DUAS LINHAS ---
    cor_secundaria_hex = negocio.cor_secundaria
    cor_secundaria_hsl = hex_to_hsl_string(cor_secundaria_hex)
    # --- FIM DA ADIÇÃO ---

    data = {
        'nome_negocio': negocio.nome_negocio,
        'tagline': negocio.tagline,
        'cor_primaria_hex': cor_primaria_hex,
        'cor_primaria_hsl': cor_primaria_hsl,

        # --- ADICIONE ESTAS DUAS LINHAS ---
        'cor_secundaria_hex': cor_secundaria_hex,
        'cor_secundaria_hsl': cor_secundaria_hsl,
        'logo_url': negocio.logo.url if negocio.logo else None,
        'portfolio_url': negocio.portfolio_url,
        'endereco': negocio.endereco,
        'telefone_contato': negocio.telefone_contato,
    }
    return JsonResponse(data)


@user_passes_test(is_admin)
def api_gestao_configuracoes(request):
    """
    API para o empreendedor logado LER e ATUALIZAR
    as configurações do seu próprio Negócio via FormData.
    """
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    # --- LER (GET) ---
    if request.method == 'GET':
        data = {
            'nome_negocio': negocio.nome_negocio,
            'tagline': negocio.tagline,
            'cor_primaria': negocio.cor_primaria,
            'cor_secundaria': negocio.cor_secundaria,
            'logo_url': negocio.logo.url if negocio.logo else None,  # <-- Adicionado
            'dias_antecedencia_maxima': negocio.dias_antecedencia_maxima,
            'portfolio_url': negocio.portfolio_url,
            'pagamento_online_habilitado': negocio.pagamento_online_habilitado,
            'endereco': negocio.endereco,
            'telefone_contato': negocio.telefone_contato,
        }
        return JsonResponse(data)

    # --- ATUALIZAR (POST) ---
    if request.method == 'POST':
        # Dados de formulário (FormData) vêm em request.POST
        try:
            negocio.nome_negocio = request.POST.get(
                'nome_negocio', negocio.nome_negocio)
            negocio.tagline = request.POST.get('tagline', negocio.tagline)
            negocio.cor_primaria = request.POST.get(
                'cor_primaria', negocio.cor_primaria)
            negocio.cor_secundaria = request.POST.get(
                'cor_secundaria', negocio.cor_secundaria)

            # --- INÍCIO DA LÓGICA DE FORMATAÇÃO DO LINK ---

            # 1. Pega o link bruto que o usuário colou
            raw_url = request.POST.get('portfolio_url', '').strip()

            if not raw_url:
                # Se o campo estiver vazio, salva como vazio/nulo
                negocio.portfolio_url = None
            elif 'canva.com' in raw_url and '/view' in raw_url:
                # 2. É um link do Canva. Vamos garantir que termine com '?embed'.

                if '/view?embed' in raw_url:
                    # 2a. O link já está perfeito. Usa ele.
                    negocio.portfolio_url = raw_url
                else:
                    # 2b. O link é (.../view) ou (.../view?utm=...).
                    # Limpa qualquer parâmetro (como ?utm=) e força o ?embed.

                    # Pega a URL base antes de qualquer '?'
                    base_url = raw_url.split('?')[0]

                    # Garante que a URL base termine exatamente com '/view'
                    if not base_url.endswith('/view'):
                        # Lida com casos como .../view/pagina-2
                        base_url = base_url.split('/view')[0] + '/view'

                    # 3. Monta o link final e correto
                    negocio.portfolio_url = base_url + '?embed'
            else:
                # 4. Não é um link do Canva ou não é um link de "view". Salva como está.
                negocio.portfolio_url = raw_url

            # --- FIM DA LÓGICA DE FORMATAÇÃO DO LINK ---

            negocio.endereco = request.POST.get('endereco', '').strip() or None
            negocio.telefone_contato = request.POST.get(
                'telefone_contato', '').strip() or None

            negocio.dias_antecedencia_maxima = request.POST.get(
                'dias_antecedencia_maxima', negocio.dias_antecedencia_maxima)

            # O 'request.POST.get' para um checkbox retorna 'true' ou 'false' como string
            pagamento_habilitado_str = request.POST.get(
                'pagamento_online_habilitado', 'false')
            negocio.pagamento_online_habilitado = (
                pagamento_habilitado_str == 'true')

            # O upload do arquivo vem em request.FILES
            if 'logo' in request.FILES:
                negocio.logo = request.FILES['logo']

            # O models.py cuida de atualizar o slug
            negocio.save()
            return JsonResponse({
                'status': 'success',
                'message': 'Configurações salvas com sucesso!',
                'new_logo_url': negocio.logo.url if negocio.logo else None
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ---
# 1. API PÚBLICA (PARA O CLIENTE LER OS AVISOS)
# ---


def api_get_avisos(request, empreendedor_slug):
    negocio = get_object_or_404(Negocio, slug=empreendedor_slug)
    avisos = Aviso.objects.filter(negocio=negocio)
    data = [
        {
            'id': aviso.id,
            'titulo': aviso.titulo,
            'conteudo': aviso.conteudo
        } for aviso in avisos
    ]
    return JsonResponse(data, safe=False)

# ---
# 2. API DO DASHBOARD (LISTAR E CRIAR AVISOS)
# ---


@user_passes_test(is_admin)
def api_gestao_avisos(request):
    try:
        negocio = request.user.empreendedor_profile.negocio
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    # --- LER (GET) ---
    if request.method == 'GET':
        avisos = Aviso.objects.filter(negocio=negocio)
        data = [
            {
                'id': aviso.id,
                'titulo': aviso.titulo,
                'conteudo': aviso.conteudo,
                'ordem': aviso.ordem
            } for aviso in avisos
        ]
        return JsonResponse(data, safe=False)

    # --- CRIAR (POST) ---
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            aviso = Aviso.objects.create(
                negocio=negocio,
                titulo=data['titulo'],
                conteudo=data['conteudo'],
                ordem=data.get('ordem', 0)
            )
            return JsonResponse({'status': 'success', 'message': 'Aviso criado!', 'id': aviso.id}, status=201)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

# ---
# 3. API DO DASHBOARD (EDITAR E EXCLUIR UM AVISO)
# ---


@user_passes_test(is_admin)
def api_gestao_aviso_detalhe(request, aviso_id):
    try:
        negocio = request.user.empreendedor_profile.negocio
        aviso = get_object_or_404(Aviso, id=aviso_id, negocio=negocio)
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Aviso.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Aviso não encontrado.'}, status=404)

    # --- INÍCIO DA CORREÇÃO (BLOCO GET FALTANTE) ---
    # --- LER DETALHE (GET) ---
    if request.method == 'GET':
        data = {
            'id': aviso.id,
            'titulo': aviso.titulo,
            'conteudo': aviso.conteudo,
            'ordem': aviso.ordem
        }
        return JsonResponse(data)
    # --- FIM DA CORREÇÃO ---

    # --- EDITAR (POST) ---
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            aviso.titulo = data.get('titulo', aviso.titulo)
            aviso.conteudo = data.get('conteudo', aviso.conteudo)
            aviso.ordem = data.get('ordem', aviso.ordem)
            aviso.save()
            return JsonResponse({'status': 'success', 'message': 'Aviso atualizado.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # --- EXCLUIR (DELETE) ---
    if request.method == 'DELETE':
        try:
            aviso.delete()
            return JsonResponse({'status': 'success', 'message': 'Aviso excluído.'}, status=204)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # Retorno para outros métodos não permitidos
    return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)


@user_passes_test(is_admin)
def api_gestao_dias_bloqueados(request):
    """
    API para LER todos os dias bloqueados e CRIAR um novo.
    Refere-se sempre ao profissional LOGADO.
    """
    try:
        profissional = request.user.empreendedor_profile
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)

    # --- LER (GET) ---
    if request.method == 'GET':
        # Retorna apenas bloqueios futuros
        hoje = timezone.now().date()
        bloqueios = DiaBloqueado.objects.filter(
            empreendedor=profissional, data__gte=hoje)
        data = [
            {
                'id': b.id,
                'data': b.data.isoformat(),
                'descricao': b.descricao,
            } for b in bloqueios
        ]
        return JsonResponse(data, safe=False)

    # --- CRIAR (POST) ---
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            data_bloqueio_str = data['data']
            data_bloqueio = datetime.strptime(
                data_bloqueio_str, '%Y-%m-%d').date()

            # --- INÍCIO DA VALIDAÇÃO (A SUA SUGESTÃO) ---
            # 1. Verifica se já existem agendamentos para este profissional neste dia
            agendamentos_existentes = Agendamento.objects.filter(
                empreendedor_executor=profissional,
                data=data_bloqueio
            )

            if agendamentos_existentes.exists():
                # 2. Se existirem, impede o bloqueio e envia a notificação
                count = agendamentos_existentes.count()
                msg = f'Não é possível bloquear este dia. Você já tem {count} agendamento(s) marcados.'
                # 400 Bad Request
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            # --- FIM DA VALIDAÇÃO ---

            # 3. Se estiver livre, cria o bloqueio
            bloqueio = DiaBloqueado.objects.create(
                empreendedor=profissional,
                data=data_bloqueio,
                descricao=data.get('descricao', 'Dia bloqueado')
            )
            return JsonResponse({'status': 'success', 'message': 'Dia bloqueado!', 'id': bloqueio.id}, status=201)

        except Exception as e:
            if 'UNIQUE constraint' in str(e):
                return JsonResponse({'status': 'error', 'message': 'Este dia já está bloqueado.'}, status=400)
            return JsonResponse({'status': 'error', 'message': f'Erro ao salvar: {str(e)}'}, status=400)


@user_passes_test(is_admin)
def api_gestao_dia_bloqueado_detalhe(request, bloqueio_id):
    """
    API para EXCLUIR um dia bloqueado.
    """
    try:
        profissional = request.user.empreendedor_profile
        bloqueio = get_object_or_404(
            DiaBloqueado, id=bloqueio_id, empreendedor=profissional)
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'Bloqueio não encontrado.'}, status=404)

    # --- EXCLUIR (DELETE) ---
    if request.method == 'DELETE':
        bloqueio.delete()
        return JsonResponse({'status': 'success', 'message': 'Bloqueio removido.'}, status=204)


@login_required(login_url=None)
@transaction.atomic  # Garante que as atualizações no User e Cliente ocorram juntas
def api_manage_profile(request, empreendedor_slug):
    """
    API para o cliente logado GERENCIAR seus próprios dados.
    """
    try:
        # Pega o usuário e o cliente logado
        user = request.user
        cliente = user.cliente
        negocio = get_object_or_404(Negocio, slug=empreendedor_slug)

        # Garante que o cliente pertence ao negócio que está acessando
        if cliente.negocio != negocio:
            return JsonResponse({'status': 'error', 'message': 'Permissão negada.'}, status=403)

    except Cliente.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Cliente não encontrado.'}, status=404)

    if request.method == 'GET':
        # --- LER DADOS ---
        return JsonResponse({
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'phone': cliente.telefone,
            'nascimento': cliente.data_nascimento
        })

    if request.method == 'POST':
        # --- ATUALIZAR DADOS ---
        try:
            data = json.loads(request.body)

            new_email = data.get('email')
            new_phone = data.get('phone')

            # Validação 1: Email (que também é o username)
            if new_email and new_email != user.email:
                if User.objects.filter(username=new_email).exclude(pk=user.pk).exists():
                    raise ValidationError(
                        'Este email já está em uso por outra conta.')
                user.email = new_email
                user.username = new_email  # Atualiza o username junto

            # Validação 2: Telefone (deve ser único NO NEGÓCIO)
            if new_phone and new_phone != cliente.telefone:
                if Cliente.objects.filter(negocio=negocio, telefone=new_phone).exclude(pk=cliente.pk).exists():
                    raise ValidationError(
                        'Este telefone já está em uso por outro cliente neste negócio.')
                cliente.telefone = new_phone

            # Atualiza os outros campos
            user.first_name = data.get('first_name', user.first_name)
            user.last_name = data.get('last_name', user.last_name)
            cliente.data_nascimento = data.get(
                'nascimento', cliente.data_nascimento)

            user.save()
            cliente.save()

            return JsonResponse({'status': 'success', 'message': 'Perfil atualizado com sucesso!'})

        except ValidationError as e:
            return JsonResponse({'status': 'error', 'message': e.message}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': f'Ocorreu um erro: {str(e)}'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Método inválido.'}, status=405)


@user_passes_test(is_admin)
@transaction.atomic
def api_gestao_equipe_detalhe(request, membro_id):
    """
    API para gerenciar um membro específico da equipe (Editar, Deletar, Ler).
    """
    try:
        negocio = request.user.empreendedor_profile.negocio
        membro = get_object_or_404(
            EmpreendedorProfile, id=membro_id, negocio=negocio)
        user = membro.user
    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Membro não encontrado.'}, status=404)

    # --- LER (GET) ---
    if request.method == 'GET':
        data = {
            'id': membro.id,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'telefone': membro.telefone,
            'foto_url': membro.foto.url if membro.foto else None,
            # Retorna se tem token configurado, mas NÃO o token real (segurança)
            'has_mp_token': bool(membro.get_access_token())
        }
        return JsonResponse(data)

    # --- EDITAR (POST) ---
    if request.method == 'POST':
        try:
            # Dados vêm de FormData (request.POST)
            user.first_name = request.POST.get('nome', user.first_name)
            user.last_name = request.POST.get('sobrenome', user.last_name)
            user.save()

            membro.telefone = request.POST.get('telefone', membro.telefone)

            # --- NOVA ADIÇÃO: Salvar Token MP ---
            mp_token = request.POST.get('mp_token')
            if mp_token and mp_token.strip():
                # Se enviou algo, criptografa e salva
                membro.set_access_token(mp_token.strip())
            # Se vier vazio, não faz nada (mantém o antigo).
            # Para limpar, você teria que implementar uma lógica específica se necessário.

            if 'foto' in request.FILES:
                membro.foto = request.FILES['foto']

            # Atualiza a senha (se uma nova foi enviada)
            nova_senha = request.POST.get('password', None)
            if nova_senha and nova_senha.strip():
                user.set_password(nova_senha)
                user.save()

            membro.save()

            return JsonResponse({
                'status': 'success',
                'message': 'Membro atualizado com sucesso!',
                'new_foto_url': membro.foto.url if membro.foto else None
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    # --- EXCLUIR (DELETE) ---
    if request.method == 'DELETE':
        try:
            # TRAVA DE SEGURANÇA: Verifica se há agendamentos futuros
            hoje = timezone.now().date()
            agendamentos_futuros = Agendamento.objects.filter(
                empreendedor_executor=membro,
                data__gte=hoje,
                status__in=['Confirmado', 'Pendente']
            ).exists()

            if agendamentos_futuros:
                return JsonResponse({'status': 'error', 'message': 'Não é possível excluir este membro, pois ele possui agendamentos futuros.'}, status=400)

            # Se não tiver, exclui o usuário (o Perfil é deletado em cascata)
            user.delete()
            return JsonResponse({'status': 'success', 'message': 'Membro excluído com sucesso.'}, status=200)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Método não permitido.'}, status=405)


@user_passes_test(is_admin)
def api_admin_get_form_data(request):
    """
    Busca todos os dados necessários (clientes, serviços, profissionais)
    para os modais de criação/edição no calendário do admin.
    """
    try:
        negocio = request.user.empreendedor_profile.negocio

        # 1. Buscar Clientes
        clientes = Cliente.objects.filter(
            negocio=negocio).select_related('user')
        clientes_data = [
            {
                'id': c.id,
                'nome': c.user.get_full_name() or c.user.username,
                'telefone': c.telefone
            } for c in clientes
        ]

        # 2. Buscar Profissionais (Equipe)
        profissionais = EmpreendedorProfile.objects.filter(
            negocio=negocio).select_related('user')
        profissionais_data = [
            {
                'id': p.id,
                'nome': p.user.get_full_name() or p.user.username
            } for p in profissionais
        ]

        # 3. Buscar Serviços e Tiers (Manutenções)
        servicos = Servico.objects.filter(
            negocio=negocio).prefetch_related('precos_manutencao')
        servicos_data = []
        for s in servicos:
            # Adiciona o serviço principal
            servicos_data.append({
                'id': f'service_{s.id}',  # ID único (ex: "service_1")
                'nome': s.nome,
                'duracao': s.duracao_minutos,
                'preco': s.preco,
                'profissionais_ids': [p.id for p in s.profissionais_que_executam.all()]
            })
            # Adiciona os tiers de manutenção
            for tier in s.precos_manutencao.all():
                servicos_data.append({
                    'id': f'tier_{tier.id}',  # ID único (ex: "tier_5")
                    'nome': f"{s.nome} - {tier.nome_tier}",
                    'duracao': tier.duracao_minutos,
                    'preco': tier.preco,
                    'profissionais_ids': [p.id for p in s.profissionais_que_executam.all()]
                })

        return JsonResponse({
            'clientes': clientes_data,
            'profissionais': profissionais_data,
            'servicos_e_tiers': servicos_data
        })

    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil não encontrado.'}, status=403)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@user_passes_test(is_admin)
@transaction.atomic
def api_admin_criar_agendamento(request):
    """
    Cria um novo agendamento (e opcionalmente um novo cliente) 
    pelo dashboard do admin.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)

    try:
        negocio = request.user.empreendedor_profile.negocio
        data = json.loads(request.body)

        cliente_id = data.get('cliente_id')
        cliente = None

        if cliente_id == 'new':
            # --- Cria um NOVO Cliente ---
            novo_email = data.get('novo_cliente_email')
            # Garante que email vazio vire None para não dar erro de duplicidade em string vazia
            if not novo_email:
                novo_email = None

            novo_telefone = re.sub(
                r'\D', '', data.get('novo_cliente_telefone', ''))

            # Validação: Email não é mais obrigatório aqui
            if not novo_telefone or not data.get('novo_cliente_nome') or not data.get('novo_cliente_nascimento'):
                raise ValidationError(
                    "Para novos clientes, Nome, Telefone e Nascimento são obrigatórios.")

            # Validação de Telefone (Username)
            if User.objects.filter(username=novo_telefone).exists():
                raise ValidationError(
                    f"Já existe um usuário com o telefone/username '{novo_telefone}'.")

            # Validação de Email (Só se foi preenchido)
            if novo_email and User.objects.filter(email=novo_email).exists():
                raise ValidationError(
                    f"O email '{novo_email}' já está em uso.")

            # Verifica duplicidade no Cliente do negócio
            if Cliente.objects.filter(negocio=negocio, telefone=novo_telefone).exists():
                raise ValidationError(
                    f"O telefone '{novo_telefone}' já está cadastrado neste negócio.")

            # Cria o User (USANDO TELEFONE COMO USERNAME)
            novo_user = User.objects.create_user(
                username=novo_telefone,  # <-- Importante: Username é o telefone
                email=novo_email,        # Pode ser None
                password=get_random_string(length=14),
                first_name=data.get('novo_cliente_nome'),
                last_name=data.get('novo_cliente_sobrenome', '')
            )

            # Cria o Cliente
            cliente = Cliente.objects.create(
                user=novo_user,
                negocio=negocio,
                telefone=novo_telefone,
                data_nascimento=data.get('novo_cliente_nascimento')
            )
        else:
            # --- Usa um Cliente Existente ---
            cliente = get_object_or_404(
                Cliente, id=int(cliente_id), negocio=negocio)

        # --- Lógica do Serviço/Tier ---
        servico_tier_id_str = data.get('servico_tier_id')
        servico = None
        tier = None

        if servico_tier_id_str.startswith('service_'):
            servico = get_object_or_404(Servico, id=int(
                servico_tier_id_str.split('_')[1]))
        elif servico_tier_id_str.startswith('tier_'):
            tier = get_object_or_404(PrecoManutencao, id=int(
                servico_tier_id_str.split('_')[1]))
            servico = tier.servico_pai

        if not servico:
            raise ValidationError("Serviço ou Manutenção inválido.")

        profissional = get_object_or_404(EmpreendedorProfile, id=int(
            data.get('profissional_id')), negocio=negocio)

        # --- Cria o Agendamento ---
        novo_agendamento = Agendamento(
            cliente=cliente,
            servico=servico,
            tier_manutencao=tier,
            empreendedor_executor=profissional,
            data=data.get('data'),
            horario=data.get('horario'),
            status='Pendente',  # <-- MUDANÇA (de 'Confirmado' para 'Pendente')
            status_pagamento='Pendente',  # Cliente paga no local
            observacoes=data.get('observacoes', '')
        )
        # O .save() vai calcular preco_final e duracao_final
        novo_agendamento.save()

        return JsonResponse({'status': 'success', 'message': 'Agendamento criado com sucesso!'}, status=201)

    except EmpreendedorProfile.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Perfil de admin não encontrado.'}, status=403)
    except (Cliente.DoesNotExist, Servico.DoesNotExist, PrecoManutencao.DoesNotExist, EmpreendedorProfile.DoesNotExist, ValidationError) as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
    except Exception as e:
        logger.error(
            f"Erro inesperado em api_admin_criar_agendamento: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Erro interno: {e}'}, status=500)


@user_passes_test(is_admin)
@transaction.atomic
def api_admin_atualizar_agendamento(request, agendamento_id):
    """
    Atualiza TODOS os dados de um agendamento (serviço, data, hora, status, etc.)
    (Substitui e expande a antiga 'api_atualizar_pagamento')
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)

    try:
        negocio = request.user.empreendedor_profile.negocio
        agendamento = get_object_or_404(
            Agendamento, id=agendamento_id, servico__negocio=negocio)

        data = json.loads(request.body)

        # Atualiza campos simples (status, pagamento, observações)
        agendamento.status = data.get('status', agendamento.status)
        agendamento.status_pagamento = data.get(
            'status_pagamento', agendamento.status_pagamento)
        agendamento.observacoes = data.get(
            'observacoes', agendamento.observacoes)

        # Atualiza campos complexos (cliente, profissional, data, hora)
        if 'cliente_id' in data:
            agendamento.cliente = get_object_or_404(
                Cliente, id=int(data['cliente_id']), negocio=negocio)

        if 'profissional_id' in data:
            agendamento.empreendedor_executor = get_object_or_404(
                EmpreendedorProfile, id=int(data['profissional_id']), negocio=negocio)

        if 'data' in data:
            agendamento.data = data['data']

        if 'horario' in data:
            agendamento.horario = data['horario']

        # Lógica de atualização do Serviço/Tier (RECALCULA PREÇO/DURAÇÃO)
        if 'servico_tier_id' in data:
            servico_tier_id_str = data.get('servico_tier_id')
            servico_novo = None
            tier_novo = None

            if servico_tier_id_str.startswith('service_'):
                servico_novo = get_object_or_404(
                    Servico, id=int(servico_tier_id_str.split('_')[1]))
            elif servico_tier_id_str.startswith('tier_'):
                tier_novo = get_object_or_404(
                    PrecoManutencao, id=int(servico_tier_id_str.split('_')[1]))
                servico_novo = tier_novo.servico_pai

            agendamento.servico = servico_novo
            agendamento.tier_manutencao = tier_novo

            # Força o recálculo do preço e duração
            agendamento.preco_final = None
            agendamento.duracao_final = None
            # O .save() irá preenchê-los com os novos valores
            # (Não mexemos no 'valor_adiantamento' de um agendamento existente)

        agendamento.save()
        return JsonResponse({'status': 'success', 'message': 'Agendamento atualizado com sucesso'})

    except (EmpreendedorProfile.DoesNotExist, Cliente.DoesNotExist, Servico.DoesNotExist, PrecoManutencao.DoesNotExist, ValidationError) as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
    except Exception as e:
        logger.error(
            f"Erro inesperado em api_admin_atualizar_agendamento: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Erro interno: {e}'}, status=500)


@user_passes_test(is_admin)
def api_admin_atualizar_horario_agendamento(request, agendamento_id):
    """
    Atualização rápida para o drag-and-drop (remanejamento).
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)

    try:
        negocio = request.user.empreendedor_profile.negocio
        agendamento = get_object_or_404(
            Agendamento, id=agendamento_id, servico__negocio=negocio)

        data = json.loads(request.body)

        # Espera "2025-11-21T14:30:00.000Z"
        nova_data_hora = data.get('start_iso')
        if not nova_data_hora:

            raise ValidationError("Nova data/hora não fornecida.")

        # --- INÍCIO DA CORREÇÃO ---

        # 1. Substitui 'Z' (Zulu/UTC) por '+00:00', que o fromisoformat entende
        if 'Z' in nova_data_hora:
            nova_data_hora = nova_data_hora.replace('Z', '+00:00')

        # 2. Cria um objeto datetime "ciente" do fuso horário (em UTC)
        dt_obj_utc = datetime.fromisoformat(nova_data_hora)

        # 3. Converte o horário UTC para o fuso horário local do servidor (definido em settings.py)
        dt_obj_local = dt_obj_utc.astimezone(timezone.get_current_timezone())

        # (Idealmente, aqui você verificaria conflitos de horário antes de salvar)

        # 4. Salva a data e hora LOCAL convertida
        agendamento.data = dt_obj_local.date()
        agendamento.horario = dt_obj_local.time()
        # --- FIM DA CORREÇÃO ---

        agendamento.save()

        return JsonResponse({'status': 'success', 'message': 'Agendamento remanejado com sucesso'})

    except (EmpreendedorProfile.DoesNotExist, Agendamento.DoesNotExist, ValidationError) as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
    except Exception as e:
        logger.error(
            f"Erro inesperado em api_admin_atualizar_horario: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Erro interno: {e}'}, status=500)


@user_passes_test(is_admin)
def api_admin_deletar_agendamento(request, agendamento_id):
    """
    Exclui um agendamento pelo dashboard do admin.
    """
    if request.method != 'DELETE':
        return JsonResponse({'status': 'error', 'message': 'Método não permitido'}, status=405)

    try:
        negocio = request.user.empreendedor_profile.negocio
        agendamento = get_object_or_404(
            Agendamento, id=agendamento_id, servico__negocio=negocio)

        agendamento.delete()

        return JsonResponse({'status': 'success', 'message': 'Agendamento excluído com sucesso'})

    except (EmpreendedorProfile.DoesNotExist, Agendamento.DoesNotExist) as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=404)
    except Exception as e:
        logger.error(
            f"Erro inesperado em api_admin_deletar_agendamento: {e}", exc_info=True)
        return JsonResponse({'status': 'error', 'message': f'Erro interno: {e}'}, status=500)


def processar_despesas_recorrentes(negocio, data_limite_solicitada=None):
    """
    Verifica despesas recorrentes e cria as instâncias de Despesa 
    até a data limite solicitada. Garante que se o usuário filtrar 
    daqui a 5 meses, as despesas sejam geradas até lá.
    """
    hoje = timezone.now().date()

    # Se nenhuma data for passada, garante pelo menos até o fim do próximo mês
    if not data_limite_solicitada:
        proximo_mes = hoje.replace(day=28) + timedelta(days=4)
        data_limite_solicitada = proximo_mes.replace(
            day=calendar.monthrange(proximo_mes.year, proximo_mes.month)[1])

    # Margem de segurança: sempre processa até o final do mês da data solicitada
    # Ex: Se pediu até 15/05, gera até 31/05 para garantir
    ano_limite = data_limite_solicitada.year
    mes_limite = data_limite_solicitada.month
    ultimo_dia_limite = calendar.monthrange(ano_limite, mes_limite)[1]
    data_limite_real = datetime(
        ano_limite, mes_limite, ultimo_dia_limite).date()

    recorrencias = DespesaRecorrente.objects.filter(negocio=negocio)

    with transaction.atomic():
        for rec in recorrencias:
            # Se a recorrência já tem data fim definida e ela é anterior à última geração, ignora
            if rec.data_fim and rec.data_fim <= rec.ultima_geracao:
                continue

            # Começa a verificar a partir do mês seguinte à última geração
            data_base = rec.ultima_geracao

            # Loop de segurança para evitar infinito (limite de 5 anos no futuro)
            safety_counter = 0

            while True:
                safety_counter += 1
                if safety_counter > 60:
                    break

                # Avança para o próximo mês de forma robusta
                ano_prox = data_base.year + (1 if data_base.month == 12 else 0)
                mes_prox = 1 if data_base.month == 12 else data_base.month + 1

                # Define a data alvo baseada no dia de vencimento original
                max_dia = calendar.monthrange(ano_prox, mes_prox)[1]
                dia_vencimento = min(rec.dia_vencimento, max_dia)

                proxima_data = datetime(
                    ano_prox, mes_prox, dia_vencimento).date()

                # CRITÉRIOS DE PARADA:

                # 1. Se a próxima despesa for DEPOIS do limite que precisamos ver
                if proxima_data > data_limite_real:
                    break

                # 2. Se a recorrência tem data fim e a próxima data passa dela
                if rec.data_fim and proxima_data > rec.data_fim:
                    break

                # Cria a despesa
                Despesa.objects.create(
                    negocio=negocio,
                    descricao=f"{rec.descricao} (Recorrente)",
                    valor=rec.valor,
                    data=proxima_data,
                    categoria=rec.categoria,
                    pago=False
                )

                # Atualiza o controle do loop e do objeto
                data_base = proxima_data
                rec.ultima_geracao = proxima_data
                rec.save()

# --- FUNÇÕES DE EXPORTAÇÃO (EXCEL E PDF) ---


@user_passes_test(is_admin)
def exportar_relatorio_excel(request):
    try:
        perfil = request.user.empreendedor_profile
        negocio = perfil.negocio
    except EmpreendedorProfile.DoesNotExist:
        return HttpResponse("Perfil não encontrado", status=403)

    tipo_relatorio = request.GET.get('tipo', 'faturamento')
    periodo = request.GET.get('periodo', 'mes')

    # --- Lógica de Datas ---
    hoje = timezone.now().date()
    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:  # Custom
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            inicio = hoje.replace(day=1)
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            fim = hoje.replace(day=ultimo_dia)

    # --- Configuração do Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Relatório {tipo_relatorio.capitalize()}"

    # Estilos
    # Remove o # da cor hexadecimal para o openpyxl
    cor_hex = negocio.cor_primaria.replace('#', '')
    font_titulo = Font(size=14, bold=True, color="FFFFFF")
    fill_titulo = PatternFill(
        start_color=cor_hex, end_color=cor_hex, fill_type="solid")
    font_header = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")

    # --- Cabeçalho do Estabelecimento ---
    ws.merge_cells('A1:E1')
    cell_titulo = ws['A1']
    cell_titulo.value = f"{negocio.nome_negocio} - Relatório de {tipo_relatorio.capitalize()}"
    cell_titulo.font = font_titulo
    cell_titulo.fill = fill_titulo
    cell_titulo.alignment = alignment_center

    # --- Metadados da Extração ---
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Gerado por: {request.user.get_full_name()} em {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = alignment_center

    ws.merge_cells('A3:E3')
    ws['A3'] = f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    ws['A3'].alignment = alignment_center

    # --- Lógica por Tipo de Relatório ---
    row_num = 5

    if tipo_relatorio == 'despesas':
        # Headers
        headers = ['Data', 'Descrição', 'Categoria', 'Valor (R$)', 'Status']
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = font_header

        # Dados
        despesas = Despesa.objects.filter(negocio=negocio, data__range=[
                                          inicio, fim]).order_by('data')
        for despesa in despesas:
            row_num += 1
            ws.cell(row=row_num, column=1).value = despesa.data.strftime(
                '%d/%m/%Y')
            ws.cell(row=row_num, column=2).value = despesa.descricao
            ws.cell(row=row_num, column=3).value = despesa.categoria
            ws.cell(row=row_num, column=4).value = float(despesa.valor)
            ws.cell(row=row_num,
                    column=5).value = "Pago" if despesa.pago else "Pendente"

    elif tipo_relatorio == 'servicos':
        # Headers
        headers = ['Serviço', 'Quantidade', 'Valor Total (R$)', 'Categoria']
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = font_header

        # Dados (Agrupados)
        servicos = Agendamento.objects.filter(
            servico__negocio=negocio,
            data__range=[inicio, fim],
            status__in=['Confirmado', 'Concluído']
        ).values('servico__nome', 'servico__categoria__nome').annotate(
            qtd=Count('id'),
            total=Sum('preco_final')
        ).order_by('-total')

        for item in servicos:
            row_num += 1
            ws.cell(row=row_num, column=1).value = item['servico__nome']
            ws.cell(row=row_num, column=2).value = item['qtd']
            ws.cell(row=row_num, column=3).value = float(item['total'] or 0)
            ws.cell(row=row_num,
                    column=4).value = item['servico__categoria__nome'] or '-'

    else:  # Faturamento (Padrão)
        # Headers
        headers = ['Data', 'Horário', 'Cliente',
                   'Serviço', 'Valor (R$)', 'Status Pagamento']
        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = font_header

        # Dados
        agendamentos = Agendamento.objects.filter(
            servico__negocio=negocio,
            data__range=[inicio, fim]
        ).order_by('data', 'horario')

        for ag in agendamentos:
            row_num += 1
            ws.cell(row=row_num, column=1).value = ag.data.strftime('%d/%m/%Y')
            ws.cell(row=row_num, column=2).value = ag.horario.strftime('%H:%M')
            ws.cell(row=row_num, column=3).value = ag.cliente.user.get_full_name()
            ws.cell(row=row_num, column=4).value = ag.servico.nome
            ws.cell(row=row_num, column=5).value = float(ag.preco_final or 0)
            ws.cell(row=row_num, column=6).value = ag.status_pagamento

    # --- Ajustar largura das colunas (CORREÇÃO APLICADA) ---
    # Iteramos enumerando as colunas (1, 2, 3...) para usar get_column_letter
    # isso evita o erro com MergedCell no cabeçalho.
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        # Obtém 'A', 'B', etc. pelo índice
        column_letter = get_column_letter(i)

        for cell in col:
            try:
                # Ignoramos células vazias ou mescladas problemáticas
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass

        # Define uma largura mínima razoável e um buffer
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=relatorio_{tipo_relatorio}_{inicio}_{fim}.xlsx'
    wb.save(response)
    return response


@user_passes_test(is_admin)
def exportar_relatorio_pdf(request):
    try:
        perfil = request.user.empreendedor_profile
        negocio = perfil.negocio
    except EmpreendedorProfile.DoesNotExist:
        return HttpResponse("Perfil não encontrado", status=403)

    tipo_relatorio = request.GET.get('tipo', 'faturamento')
    periodo = request.GET.get('periodo', 'mes')

    # --- Lógica de Datas (Mesma do Excel) ---
    hoje = timezone.now().date()
    if periodo == 'semana':
        inicio = hoje - timedelta(days=hoje.weekday())
        fim = inicio + timedelta(days=6)
    elif periodo == 'mes':
        inicio = hoje.replace(day=1)
        ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
        fim = hoje.replace(day=ultimo_dia)
    elif periodo == 'ano':
        inicio = hoje.replace(month=1, day=1)
        fim = hoje.replace(month=12, day=31)
    else:
        data_inicio = request.GET.get('inicio')
        data_fim = request.GET.get('fim')
        if data_inicio and data_fim:
            inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
            fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
        else:
            inicio = hoje.replace(day=1)
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            fim = hoje.replace(day=ultimo_dia)

    # --- Preparação dos Dados para o Template ---
    context = {
        'negocio': negocio,
        'usuario': request.user,
        'data_geracao': timezone.now(),
        'periodo_inicio': inicio,
        'periodo_fim': fim,
        'tipo_relatorio': tipo_relatorio.capitalize(),
        'logo_url': request.build_absolute_uri(negocio.logo.url) if negocio.logo else None,
        'cor_primaria': negocio.cor_primaria,
    }

    if tipo_relatorio == 'despesas':
        dados = Despesa.objects.filter(negocio=negocio, data__range=[
                                       inicio, fim]).order_by('data')
        total = dados.aggregate(Sum('valor'))['valor__sum'] or 0
        context['headers'] = ['Data', 'Descrição',
                              'Categoria', 'Valor', 'Status']
        context['rows'] = dados
        context['total_geral'] = total
        context['template_type'] = 'despesas'

    elif tipo_relatorio == 'servicos':
        dados = Agendamento.objects.filter(
            servico__negocio=negocio,
            data__range=[inicio, fim],
            status__in=['Confirmado', 'Concluído']
        ).values('servico__nome', 'servico__categoria__nome').annotate(
            qtd=Count('id'),
            total=Sum('preco_final')
        ).order_by('-total')
        total = sum(item['total'] or 0 for item in dados)
        context['headers'] = ['Serviço', 'Categoria', 'Qtd', 'Total']
        context['rows'] = dados
        context['total_geral'] = total
        context['template_type'] = 'servicos'

    else:  # Faturamento/Geral
        dados = Agendamento.objects.filter(
            servico__negocio=negocio,
            data__range=[inicio, fim]
        ).order_by('data', 'horario')

        # Filtra apenas pagos ou confirmados para o total
        total = dados.filter(status_pagamento='Pago').aggregate(
            Sum('preco_final'))['preco_final__sum'] or 0

        context['headers'] = ['Data', 'Horário',
                              'Cliente', 'Serviço', 'Valor', 'Status']
        context['rows'] = dados
        context['total_geral'] = total
        context['template_type'] = 'faturamento'

    # --- Renderização do PDF ---
    template_path = 'agendamentos/relatorios/pdf_template.html'
    template = get_template(template_path)
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="relatorio_{tipo_relatorio}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=500)
    return response


@csrf_exempt
def api_lembretes_24h(request):
    """
    API que retorna JSON com agendamentos de amanhã.
    Protegida por um Token no Header.
    """
    # 1. Segurança Básica: Verifica se quem chamou tem a senha correta
    token_recebido = request.headers.get('X-N8N-Token')
    # Defina isso no Render depois
    token_secreto = os.getenv('N8N_ACCESS_TOKEN', 'senha_super_secreta_123')

    if token_recebido != token_secreto:
        return JsonResponse({'erro': 'Acesso negado'}, status=403)

    # 2. Lógica: Buscar agendamentos de amanhã
    hoje = timezone.now().date()
    amanha = hoje + timedelta(days=1)

    agendamentos = Agendamento.objects.filter(
        status__in=['Confirmado', 'Pendente'],
        data=amanha
    ).select_related('cliente', 'cliente__user', 'servico', 'servico__negocio')

    lista_envio = []

    for agendamento in agendamentos:
        # Tratamento do Telefone
        telefone = agendamento.cliente.telefone
        telefone_limpo = re.sub(r'\D', '', telefone)
        if len(telefone_limpo) == 11:  # Adiciona 55 se for celular BR sem DDI
            telefone_limpo = f"55{telefone_limpo}"

        dados = {
            "cliente_nome": agendamento.cliente.user.get_full_name() or agendamento.cliente.user.username,
            "cliente_telefone": telefone_limpo,
            "servico_nome": agendamento.servico.nome,
            "data": agendamento.data.strftime('%d/%m/%Y'),
            "horario": agendamento.horario.strftime('%H:%M'),
            "local_nome": agendamento.servico.negocio.nome_negocio
        }
        lista_envio.append(dados)

    # Retorna a lista para o n8n
    return JsonResponse({'quantidade': len(lista_envio), 'lista': lista_envio}, safe=False)
